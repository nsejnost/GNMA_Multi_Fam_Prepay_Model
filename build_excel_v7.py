"""
V7 Excel calculator builder — multiplicative architecture.

Six tabs:
  1. README             — orientation
  2. Multiplier_Params  — fitted parameters per multiplier (named cells)
  3. Lookups            — small categorical lookup tables (program, purpose, optional issuer overlay)
  4. Loan_Snapshot      — per-loan inputs → 9 multiplier columns → product → capped SMM → CPR → log-attribution
  5. Deal_Aggregator    — UPB-weighted deal-level rollup
  6. Benchmarks         — 12-row SanCap benchmark grid (informational sanity check)

All math is via live Excel formulas; no values are hard-coded into cell values for the
math chain. Closed-form formulas for the 6 numeric multipliers (so the user can audit
the parameter-to-multiplier transform). VLOOKUP only for the 3 small categorical tables.

Math form (per the V7 plan):
    pred_SMM      = MIN(base_SMM × M_age × M_rate × ... × M_burnout, smm_cap)
    pred_CPR      = 1 - (1 - pred_SMM)^12
    base_CPR      = 1 - (1 - base_SMM)^12
    log_mult_j    = LN(M_factor_j)
    attribution_j = log_mult_j / SUM(log_mults) × (pred_CPR - base_CPR)
"""
import json, math, os, sys
from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

import v7_multipliers as v7m

HERE = Path(__file__).resolve().parent
MODEL_JSON = HERE / "model_data_v7.json"
ISSUER_JSON = HERE / "issuer_residuals.json"
SAMPLE_PARQUET = HERE / "sample_loans.parquet"
OUT_XLSX = HERE / "V7_Excel_Calculator.xlsx"


# ---------- Style helpers (match V6F conventions) ----------
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
SUBHEADER_FILL = PatternFill("solid", fgColor="374151")
PARAM_FILL = PatternFill("solid", fgColor="FEF3C7")
FORMULA_FILL = PatternFill("solid", fgColor="E0F2FE")
BASELINE_FILL = PatternFill("solid", fgColor="FFEDD5")
BENCH_FILL = PatternFill("solid", fgColor="DBEAFE")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
MONO = Font(name="Consolas")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
thin = Side(border_style="thin", color="9CA3AF")
BORDER = Border(top=thin, bottom=thin, left=thin, right=thin)


def cl(i): return get_column_letter(i)


# ===========================================================================
def main():
    if not MODEL_JSON.exists():
        sys.exit(f"ERROR: {MODEL_JSON} not found - run train_v7.py first")
    md = json.load(open(MODEL_JSON))
    issuer_data = json.load(open(ISSUER_JSON)) if ISSUER_JSON.exists() else {'issuers': []}
    params = md['multipliers']

    # Resolve sample
    panel_env = os.environ.get('GNMA_PANEL_PARQUET')
    if panel_env and Path(panel_env).exists():
        # Pull a stratified sample from the latest period of the full panel
        sample = regenerate_sample(panel_env)
    elif SAMPLE_PARQUET.exists():
        sample = pd.read_parquet(SAMPLE_PARQUET).reset_index(drop=True)
    else:
        sys.exit(f"ERROR: neither GNMA_PANEL_PARQUET nor {SAMPLE_PARQUET}")
    sample = sample[sample['loan_age_months'].notna() & sample['vintage_year'].notna()].reset_index(drop=True)

    print(f"Building {OUT_XLSX.name}  test_AUC={md['metadata']['test_auc']:.4f}  "
          f"max_pred_CPR={md['metadata']['max_cpr_full_panel']:.1f}%")
    print(f"  Loans in sample: {len(sample)}")

    wb = Workbook()
    wb.remove(wb.active)

    build_readme(wb, md)
    param_cells = build_multiplier_params(wb, params, md)
    lookup_cells = build_lookups(wb, params, issuer_data)
    build_loan_snapshot(wb, sample, params, param_cells, lookup_cells)
    build_deal_aggregator(wb, sample, params)
    build_benchmarks_tab(wb, params, md)

    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}  ({OUT_XLSX.stat().st_size / 1024:.1f} KB)")


def regenerate_sample(panel_path):
    """Same stratified-sample logic as V6E/V6F's build_excel.py."""
    print(f"  Reading panel: {panel_path}")
    df = pd.read_parquet(panel_path)
    df = df.sort_values(['loan_id', 'period'])
    df['itm'] = (df['refi_incentive_bps'] > 0).astype(int)
    df['cum_itm'] = df.groupby('loan_id')['itm'].cumsum()
    snap_period = sorted(df['period'].astype(str).unique())[-2]
    snap = df[(df['period'].astype(str) == snap_period)
              & (df['prepay_eligible'] == 1) & (df['has_valid_data'] == 1)].copy()
    buckets = []
    for issuer_pat, n in [
        ('LUMENT', 35), ('BERKADIA', 30), ('WALKER', 25), ('GREYSTONE', 25),
        ('MERCHANTS', 20), ('PNC', 20), ('WELLS FARGO', 20), ('CAPITAL FUNDING', 20),
        ('DWIGHT', 15), ('KEYBANK', 15), ('JLL', 10), ('BELLWETHER', 8),
        ('PGIM', 5), ('NORTHMARQ', 5),
    ]:
        m = snap['issuer_name'].fillna('').str.upper().str.contains(issuer_pat, na=False, regex=False)
        sub = snap[m]
        if len(sub) > 0:
            buckets.append(sub.sample(min(n, len(sub)), random_state=42))
    out = pd.concat(buckets).drop_duplicates('loan_id').reset_index(drop=True)
    out.to_parquet(SAMPLE_PARQUET)
    return out


# ---------------------------------------------------------------------------
# Sheet 1: README
# ---------------------------------------------------------------------------
def build_readme(wb, md):
    ws = wb.create_sheet("README")
    rows = [
        ("V7 Voluntary Prepayment Calculator (multiplicative)", BOLD),
        ("", None),
        (f"Test AUC: {md['metadata']['test_auc']:.4f}   "
         f"Test log-loss: {md['metadata']['test_log_loss']:.5f}   "
         f"Free params: {md['metadata']['n_free_params']}", None),
        (f"Training period: {md['metadata']['period_range'][0]} to {md['metadata']['period_range'][1]}", None),
        (f"Training events: {md['metadata']['training_events']:,} prepays in {md['metadata']['training_pop_n']:,} loan-months", None),
        (f"Max predicted CPR over full panel: {md['metadata']['max_cpr_full_panel']:.2f}%   "
         f"(SMM cap: {md['metadata']['smm_cap']:.3f})", None),
        ("", None),
        ("Architecture", BOLD),
        ("  predicted_SMM = MIN(base_SMM × M_age × M_rate × M_size × M_program", MONO),
        ("                   × M_purpose × M_lockout × M_maturity × M_burnout, smm_cap)", MONO),
        ("  M_rate input = net_refi_bps = gross_refi_bps - 12.5 × (ppp + 1)  ← penalty bp-deductible", MONO),
        ("  predicted_CPR = 1 - (1 - predicted_SMM)^12", MONO),
        ("", None),
        ("Tabs", BOLD),
        ("  Multiplier_Params - 9 multipliers + base_SMM + smm_cap, all parameters editable", None),
        ("  Lookups           - categorical multipliers (program, purpose, issuer overlay)", None),
        ("  Loan_Snapshot     - per-loan inputs, 9 multiplier columns, capped SMM, CPR, log-attribution", None),
        ("  Deal_Aggregator   - UPB-weighted deal-level rollup", None),
        ("  Benchmarks        - SanCap reference grid (informational; flags divergences > 10pp)", None),
        ("", None),
        ("Math (multiplicative form)", BOLD),
        ("  base_CPR     = 1 - (1 - base_SMM)^12", MONO),
        ("  log_mult_j   = LN(M_factor_j)", MONO),
        ("  attribution  = log_mult_j / SUM(log_mults) × (pred_CPR - base_CPR)", MONO),
        ("", None),
        ("Issuer overlay (optional)", BOLD),
        ("  USE_ISSUER_OVERLAY in Multiplier_Params: 0 (off, default) or 1 (on).", None),
        ("  When on, applies the empirical residual ratio from issuer_residuals.json as a 10th multiplier.", None),
        ("  Default 1.0 keeps the core model issuer-neutral.", None),
        ("", None),
        ("Sanity check (column 'Sanity' in Loan_Snapshot)", BOLD),
        ("  base_CPR + SUM(attribution) - pred_CPR  should be ≈ 0 (within rounding).", MONO),
    ]
    for i, (text, font) in enumerate(rows, 1):
        c = ws.cell(i, 1, text)
        if font is not None: c.font = font
    ws.column_dimensions['A'].width = 110


# ---------------------------------------------------------------------------
# Sheet 2: Multiplier_Params
# ---------------------------------------------------------------------------
def build_multiplier_params(wb, params, md):
    """Write all parameters as named cells; return a dict mapping logical name → "Multiplier_Params!$B$X" address."""
    ws = wb.create_sheet("Multiplier_Params")
    ws.cell(1, 1, "V7 Multiplier Parameters (edit at your own risk)").font = Font(bold=True, size=14)
    ws.cell(2, 1, "Each cell is referenced by name from Loan_Snapshot formulas. Editing cascades through.").font = Font(italic=True, color="6B7280")

    cells = {}
    row = 4

    def write(label, value, name, fmt="0.000000"):
        nonlocal row
        ws.cell(row, 1, label).font = MONO
        c = ws.cell(row, 2, value)
        c.font = MONO; c.fill = PARAM_FILL; c.number_format = fmt
        wb.defined_names[name] = DefinedName(name, attr_text=f"Multiplier_Params!$B${row}")
        cells[name] = f"Multiplier_Params!$B${row}"
        row += 1

    ws.cell(row, 1, "Global"); ws.cell(row, 1).font = BOLD; row += 1
    write("base_SMM",         params['base_smm'],   'V7_BASE_SMM',    "0.0000000")
    write("smm_cap",          params['smm_cap'],    'V7_SMM_CAP',     "0.0000")
    row += 1

    ws.cell(row, 1, "M_age (piecewise linear, 6 knots)"); ws.cell(row, 1).font = BOLD; row += 1
    knots_x = params['M_age']['knots_x']
    knots_y = params['M_age']['knots_y']
    for i, (x, y) in enumerate(zip(knots_x, knots_y)):
        write(f"  knot {i}: x = {x:.0f}m → y", y, f'V7_AGE_KNOT_Y_{i}', "0.0000")
    # Also store knot_x in cells for the formula to reference
    knot_x_addrs = []
    ws.cell(row, 1, "  (knot_x values, frozen)").font = MONO; row += 1
    for i, x in enumerate(knots_x):
        ws.cell(row, 1, f"    knot_x_{i}").font = MONO
        c = ws.cell(row, 2, x); c.font = MONO; c.number_format = "0"
        wb.defined_names[f'V7_AGE_KNOT_X_{i}'] = DefinedName(f'V7_AGE_KNOT_X_{i}',
                                                              attr_text=f"Multiplier_Params!$B${row}")
        cells[f'V7_AGE_KNOT_X_{i}'] = f"Multiplier_Params!$B${row}"
        row += 1
    row += 1

    ws.cell(row, 1, "M_rate (sigmoid: floor + asymp/(1+exp(-(grf-mid)/slope)))"); ws.cell(row, 1).font = BOLD; row += 1
    write("  floor",     params['M_rate']['floor'],     'V7_RATE_FLOOR',     "0.0000")
    write("  asymptote", params['M_rate']['asymptote'], 'V7_RATE_ASYMPTOTE', "0.0000")
    write("  mid (bp)",  params['M_rate']['mid'],       'V7_RATE_MID',       "0.00")
    write("  slope",     params['M_rate']['slope'],     'V7_RATE_SLOPE',     "0.0000")
    row += 1

    ws.cell(row, 1, "M_size (smooth log on log_upb)"); ws.cell(row, 1).font = BOLD; row += 1
    write("  intercept",        params['M_size']['intercept'],      'V7_SIZE_INTERCEPT',  "0.0000")
    write("  slope",            params['M_size']['slope'],          'V7_SIZE_SLOPE',      "0.000000")
    write("  log_upb_anchor",   params['M_size']['log_upb_anchor'], 'V7_SIZE_LOG_ANCHOR', "0.000000")
    write("  low (clip)",       params['M_size']['low'],            'V7_SIZE_LOW',        "0.0000")
    write("  high (clip)",      params['M_size']['high'],           'V7_SIZE_HIGH',       "0.0000")
    row += 1

    ws.cell(row, 1, "M_purpose (NC triangular hump centered at peak_age)"); ws.cell(row, 1).font = BOLD; row += 1
    write("  NC_bump",  params['M_purpose']['NC_bump'],  'V7_PURPOSE_NC_BUMP',  "0.0000")
    write("  peak_age", params['M_purpose']['peak_age'], 'V7_PURPOSE_PEAK_AGE', "0.00")
    write("  width",    params['M_purpose']['width'],    'V7_PURPOSE_WIDTH',    "0.00")
    row += 1

    ws.cell(row, 1, "M_lockout (1 + amp*exp(-msle/tau) for msle ∈ (0,12])"); ws.cell(row, 1).font = BOLD; row += 1
    write("  amplitude", params['M_lockout']['amplitude'], 'V7_LOCKOUT_AMP', "0.0000")
    write("  tau",       params['M_lockout']['tau'],       'V7_LOCKOUT_TAU', "0.0000")
    row += 1

    ws.cell(row, 1, "M_maturity (1 + amp*max(0, (cutoff-mtm)/cutoff))"); ws.cell(row, 1).font = BOLD; row += 1
    write("  amplitude", params['M_maturity']['amplitude'], 'V7_MTY_AMP',    "0.0000")
    write("  cutoff",    params['M_maturity']['cutoff'],    'V7_MTY_CUTOFF', "0.00")
    row += 1

    ws.cell(row, 1, "M_burnout (max(floor, 1 - slope*burn_ratio))"); ws.cell(row, 1).font = BOLD; row += 1
    write("  floor", params['M_burnout']['floor'], 'V7_BURN_FLOOR', "0.0000")
    write("  slope", params['M_burnout']['slope'], 'V7_BURN_SLOPE', "0.0000")
    row += 1

    ws.cell(row, 1, "Issuer overlay toggle (0=off, 1=on)"); ws.cell(row, 1).font = BOLD; row += 1
    ws.cell(row, 1, "  USE_ISSUER_OVERLAY").font = MONO
    c = ws.cell(row, 2, 0); c.font = MONO; c.fill = PARAM_FILL
    wb.defined_names['V7_USE_ISSUER'] = DefinedName('V7_USE_ISSUER', attr_text=f"Multiplier_Params!$B${row}")
    cells['V7_USE_ISSUER'] = f"Multiplier_Params!$B${row}"
    row += 1
    row += 1

    ws.cell(row, 1, "Diagnostics (read-only)").font = BOLD; row += 1
    write("  test_AUC",      md['metadata']['test_auc'],      'V7_TEST_AUC',     "0.0000")
    write("  test_log_loss", md['metadata']['test_log_loss'], 'V7_TEST_LL',      "0.00000")
    write("  base_CPR",      md['metadata']['base_cpr'] / 100, 'V7_BASE_CPR',    "0.00%")
    write("  panel_actual_CPR", md['metadata']['panel_actual_cpr'] / 100, 'V7_PANEL_ACTUAL', "0.00%")
    write("  max_pred_CPR (full panel)", md['metadata']['max_cpr_full_panel'] / 100, 'V7_MAX_CPR', "0.00%")

    ws.column_dimensions['A'].width = 56
    ws.column_dimensions['B'].width = 16
    ws.freeze_panes = "A4"
    return cells


# ---------------------------------------------------------------------------
# Sheet 3: Lookups
# ---------------------------------------------------------------------------
def build_lookups(wb, params, issuer_data):
    """Three small categorical lookup tables; returns dict of named ranges."""
    ws = wb.create_sheet("Lookups")
    ws.cell(1, 1, "Categorical multiplier lookups").font = Font(bold=True, size=14)

    lookup_cells = {}

    # Program lookup (col A=key, col B=mult)
    ws.cell(3, 1, "M_program").font = BOLD
    ws.cell(4, 1, "fha_pattern").font = WHITE; ws.cell(4, 1).fill = HEADER_FILL
    ws.cell(4, 2, "multiplier").font = WHITE; ws.cell(4, 2).fill = HEADER_FILL
    prog_keys = ['232', '538', '223a7', '223f', 'default']
    for i, k in enumerate(prog_keys):
        r = 5 + i
        ws.cell(r, 1, k).font = MONO
        c = ws.cell(r, 2, params['M_program'].get(k, 1.0))
        c.font = MONO; c.fill = PARAM_FILL; c.number_format = "0.0000"
    prog_table = f"Lookups!$A$5:$B${5 + len(prog_keys) - 1}"
    wb.defined_names['V7_PROGRAM_TABLE'] = DefinedName('V7_PROGRAM_TABLE', attr_text=prog_table)
    lookup_cells['V7_PROGRAM_TABLE'] = prog_table
    # Default cell (named for fallback formula)
    wb.defined_names['V7_PROGRAM_DEFAULT'] = DefinedName('V7_PROGRAM_DEFAULT',
        attr_text=f"Lookups!$B${5 + prog_keys.index('default')}")

    # Issuer overlay lookup (issuer_number → residual_ratio); only top issuers
    ws.cell(3, 5, "Issuer overlay (residual ratio)").font = BOLD
    ws.cell(4, 5, "issuer_number").font = WHITE; ws.cell(4, 5).fill = HEADER_FILL
    ws.cell(4, 6, "issuer_name").font = WHITE;   ws.cell(4, 6).fill = HEADER_FILL
    ws.cell(4, 7, "residual_ratio").font = WHITE; ws.cell(4, 7).fill = HEADER_FILL
    ws.cell(4, 8, "n").font = WHITE;             ws.cell(4, 8).fill = HEADER_FILL
    iss_rows = issuer_data.get('issuers', [])
    for i, irow in enumerate(iss_rows[:30]):
        r = 5 + i
        ws.cell(r, 5, str(irow['issuer_number'])).font = MONO
        ws.cell(r, 6, irow['issuer_name'])
        c = ws.cell(r, 7, irow['residual_ratio'])
        c.font = MONO; c.number_format = "0.0000"
        ws.cell(r, 8, irow['n'])
    iss_table = f"Lookups!$E$5:$G${5 + min(len(iss_rows), 30) - 1}" if iss_rows else "Lookups!$E$5:$G$5"
    wb.defined_names['V7_ISSUER_TABLE'] = DefinedName('V7_ISSUER_TABLE', attr_text=iss_table)
    lookup_cells['V7_ISSUER_TABLE'] = iss_table

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 36
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 10
    ws.freeze_panes = "A5"
    return lookup_cells


# ---------------------------------------------------------------------------
# Sheet 4: Loan_Snapshot
# ---------------------------------------------------------------------------
def build_loan_snapshot(wb, sample, params, param_cells, lookup_cells):
    ws = wb.create_sheet("Loan_Snapshot")

    # Column layout: ID(8) + INPUT(11) + DERIVED(2) + 9 mults + product + capped_smm + cpr + 9 attr + sanity = 42
    ID_COLS = ["loan_id", "deal_id", "pool_cusip", "issuer_name", "fha_category", "loan_purpose",
               "vintage_year", "period"]
    INPUT_COLS = ["upb", "loan_rate_pct", "plc_rate_bps", "prepay_penalty_points",
                  "loan_age_months", "cum_itm", "loan_maturity_yyyymm", "lockout_end_yyyymm",
                  "issuer_number", "burn_ratio_in", "is_NC_in"]
    DERIVED_COLS = ["log_upb", "gross_refi_bps", "net_refi_bps"]
    MULT_COLS = list(v7m.MULTIPLIER_NAMES) + ["M_issuer"]   # 10 (issuer is overlay)
    SUMMARY_COLS = ["product", "capped_SMM", "pred_CPR"]
    ATTR_COLS = [f"attr_{n}" for n in v7m.MULTIPLIER_NAMES]   # 9 (issuer not in attribution; it's a post-hoc overlay)

    n_id = len(ID_COLS); n_input = len(INPUT_COLS); n_derived = len(DERIVED_COLS)
    n_mult = len(MULT_COLS); n_summ = len(SUMMARY_COLS); n_attr = len(ATTR_COLS)

    COL_ID_START = 1
    COL_INPUT_START = COL_ID_START + n_id            # 9
    COL_DERIVED_START = COL_INPUT_START + n_input    # 20
    COL_MULT_START = COL_DERIVED_START + n_derived   # 22
    COL_SUMM_START = COL_MULT_START + n_mult         # 32
    COL_ATTR_START = COL_SUMM_START + n_summ         # 35
    COL_SANITY = COL_ATTR_START + n_attr             # 44

    # Column index lookups
    ID = {name: COL_ID_START + i for i, name in enumerate(ID_COLS)}
    IN = {name: COL_INPUT_START + i for i, name in enumerate(INPUT_COLS)}
    DR = {name: COL_DERIVED_START + i for i, name in enumerate(DERIVED_COLS)}
    M = {name: COL_MULT_START + i for i, name in enumerate(MULT_COLS)}
    SU = {name: COL_SUMM_START + i for i, name in enumerate(SUMMARY_COLS)}
    AT = {name: COL_ATTR_START + i for i, name in enumerate(ATTR_COLS)}

    # ----- Row 1 banners -----
    def banner(c1, c2, text, fill=SUBHEADER_FILL):
        ws.merge_cells(start_row=1, end_row=1, start_column=c1, end_column=c2)
        c = ws.cell(1, c1, text); c.fill = fill; c.font = WHITE; c.alignment = CENTER
    banner(COL_ID_START,      COL_INPUT_START - 1, "Loan Identifiers")
    banner(COL_INPUT_START,   COL_DERIVED_START - 1, "Raw Inputs (editable)", PatternFill("solid", fgColor="064E3B"))
    banner(COL_DERIVED_START, COL_MULT_START - 1, "Derived inputs (formulas)")
    banner(COL_MULT_START,    COL_SUMM_START - 1, "V7 Multipliers (formulas)", PatternFill("solid", fgColor="312E81"))
    banner(COL_SUMM_START,    COL_ATTR_START - 1, "Pred SMM/CPR", PatternFill("solid", fgColor="92400E"))
    banner(COL_ATTR_START,    COL_SANITY - 1, "Per-multiplier attribution", PatternFill("solid", fgColor="9D174D"))
    banner(COL_SANITY,        COL_SANITY, "Check", PatternFill("solid", fgColor="6B7280"))

    # ----- Row 2 headers -----
    hdrs = ID_COLS + INPUT_COLS + DERIVED_COLS + MULT_COLS + SUMMARY_COLS + ATTR_COLS + ["Sanity"]
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(2, j, h); c.fill = HEADER_FILL; c.font = WHITE; c.alignment = CENTER

    DATA_START_ROW = 3

    # ----- Per-loan rows -----
    for i, s in sample.iterrows():
        r = DATA_START_ROW + i
        # Identifiers
        ws.cell(r, ID['loan_id'], str(s['loan_id'])).font = MONO
        issuer_short = str(s['issuer_name']).split()[0].replace(',', '').upper()[:12]
        era = "POST21" if (pd.notna(s['vintage_year']) and s['vintage_year'] >= 2021) else "PRE21"
        ws.cell(r, ID['deal_id'], f"{issuer_short}_{era}")
        ws.cell(r, ID['pool_cusip'], str(s['pool_cusip'])).font = MONO
        ws.cell(r, ID['issuer_name'], str(s['issuer_name']))
        ws.cell(r, ID['fha_category'], str(s.get('fha_category', '')))
        ws.cell(r, ID['loan_purpose'], str(s.get('loan_purpose', '')))
        ws.cell(r, ID['vintage_year'], int(s['vintage_year']) if pd.notna(s['vintage_year']) else None)
        ws.cell(r, ID['period'], str(s['period']))

        # Raw inputs
        ws.cell(r, IN['upb'], float(s['upb'])).number_format = "#,##0"
        ws.cell(r, IN['loan_rate_pct'], float(s['loan_rate'])).number_format = "0.000"
        ws.cell(r, IN['plc_rate_bps'], float(s['plc_rate_bps'])).number_format = "0.0"
        ws.cell(r, IN['prepay_penalty_points'],
                float(s.get('prepay_penalty_points', 0)) if pd.notna(s.get('prepay_penalty_points', 0)) else 0).number_format = "0.0"
        ws.cell(r, IN['loan_age_months'], float(s['loan_age_months']))
        ws.cell(r, IN['cum_itm'], int(s.get('cum_itm', 0))).number_format = "0"
        # Maturity / lockout YYYYMM ints
        mat_str = str(s.get('loan_maturity_date', '')).strip()
        ws.cell(r, IN['loan_maturity_yyyymm'],
                int(mat_str[:6]) if len(mat_str) >= 6 and mat_str[:6].isdigit() else '')
        lo_str = str(s.get('lockout_end_date', '')).strip()
        ws.cell(r, IN['lockout_end_yyyymm'],
                int(lo_str[:6]) if len(lo_str) >= 6 and lo_str[:6].isdigit() else '')
        ws.cell(r, IN['issuer_number'], str(s.get('issuer_number', '')))
        # burn_ratio: pre-computed in the sample (matches V7 derivation)
        age = max(1, float(s['loan_age_months']))
        br = min(1.0, max(0.0, float(s.get('cum_itm', 0)) / age))
        ws.cell(r, IN['burn_ratio_in'], br).number_format = "0.0000"
        ws.cell(r, IN['is_NC_in'], 1 if str(s.get('loan_purpose', '')).upper() == 'NC' else 0)

        # Derived: log_upb, gross_refi_bps, net_refi_bps (penalty-deductible)
        ws.cell(r, DR['log_upb'],
                f"=LN(1+{cl(IN['upb'])}{r})").number_format = "0.000"
        ws.cell(r, DR['gross_refi_bps'],
                f"={cl(IN['loan_rate_pct'])}{r}*100-{cl(IN['plc_rate_bps'])}{r}").number_format = "0.0"
        # net_refi_bps = gross - 12.5*(ppp+1) — matches panel's `refi_incentive_bps` formula
        ws.cell(r, DR['net_refi_bps'],
                f"={cl(DR['gross_refi_bps'])}{r}-12.5*({cl(IN['prepay_penalty_points'])}{r}+1)"
                ).number_format = "0.0"

        # ----- Multipliers (closed-form formulas) -----

        # M_age: piecewise linear interpolation through 6 fixed knots
        # Build a nested IF using the named knot_x and knot_y cells
        age_col = cl(IN['loan_age_months'])
        # M_age formula: long but explicit. Use FORECAST.LINEAR-style with TREND? Simpler: nested IFs.
        f_age = (
            f"=IF({age_col}{r}<=V7_AGE_KNOT_X_0,V7_AGE_KNOT_Y_0,"
            f"IF({age_col}{r}>=V7_AGE_KNOT_X_5,V7_AGE_KNOT_Y_5,"
            f"IF({age_col}{r}<=V7_AGE_KNOT_X_1,"
                f"V7_AGE_KNOT_Y_0+({age_col}{r}-V7_AGE_KNOT_X_0)/(V7_AGE_KNOT_X_1-V7_AGE_KNOT_X_0)*(V7_AGE_KNOT_Y_1-V7_AGE_KNOT_Y_0),"
            f"IF({age_col}{r}<=V7_AGE_KNOT_X_2,"
                f"V7_AGE_KNOT_Y_1+({age_col}{r}-V7_AGE_KNOT_X_1)/(V7_AGE_KNOT_X_2-V7_AGE_KNOT_X_1)*(V7_AGE_KNOT_Y_2-V7_AGE_KNOT_Y_1),"
            f"IF({age_col}{r}<=V7_AGE_KNOT_X_3,"
                f"V7_AGE_KNOT_Y_2+({age_col}{r}-V7_AGE_KNOT_X_2)/(V7_AGE_KNOT_X_3-V7_AGE_KNOT_X_2)*(V7_AGE_KNOT_Y_3-V7_AGE_KNOT_Y_2),"
            f"IF({age_col}{r}<=V7_AGE_KNOT_X_4,"
                f"V7_AGE_KNOT_Y_3+({age_col}{r}-V7_AGE_KNOT_X_3)/(V7_AGE_KNOT_X_4-V7_AGE_KNOT_X_3)*(V7_AGE_KNOT_Y_4-V7_AGE_KNOT_Y_3),"
                f"V7_AGE_KNOT_Y_4+({age_col}{r}-V7_AGE_KNOT_X_4)/(V7_AGE_KNOT_X_5-V7_AGE_KNOT_X_4)*(V7_AGE_KNOT_Y_5-V7_AGE_KNOT_Y_4)"
            f"))))))"
        )
        ws.cell(r, M['M_age'], f_age).number_format = "0.0000"

        # M_rate: 4-param sigmoid on NET refi bps (penalty already deducted)
        net_col = cl(DR['net_refi_bps'])
        f_rate = (f"=V7_RATE_FLOOR+V7_RATE_ASYMPTOTE/"
                  f"(1+EXP(-({net_col}{r}-V7_RATE_MID)/V7_RATE_SLOPE))")
        ws.cell(r, M['M_rate'], f_rate).number_format = "0.0000"

        # M_size: clip(intercept + slope*(log_upb - anchor), low, high)
        log_col = cl(DR['log_upb'])
        f_size = (f"=MAX(V7_SIZE_LOW,MIN(V7_SIZE_HIGH,"
                  f"V7_SIZE_INTERCEPT+V7_SIZE_SLOPE*({log_col}{r}-V7_SIZE_LOG_ANCHOR)))")
        ws.cell(r, M['M_size'], f_size).number_format = "0.0000"

        # M_program: VLOOKUP(fha_category, table, 2, FALSE) with default fallback via IFERROR
        fha_col = cl(ID['fha_category'])
        # Match by substring: SUMPRODUCT trick. Simpler: try exact-match VLOOKUP, fall back to default.
        # Our fha_category strings are like "232", "223a7", "538"; lookup table uses these patterns directly.
        # If a row's fha_category contains a key as substring, the SUMPRODUCT below grabs the first matching multiplier.
        # Use INDEX(MATCH) on substring search.
        f_prog = (
            f'=IFERROR(INDEX(V7_PROGRAM_TABLE,'
            f'MATCH(TRUE,ISNUMBER(SEARCH(INDEX(V7_PROGRAM_TABLE,0,1),{fha_col}{r})),0),2),'
            f'V7_PROGRAM_DEFAULT)'
        )
        # Excel array-formula context required; openpyxl will mark with curly braces via formula_attributes.
        # For simplicity and broad compatibility, use a plain VLOOKUP by exact substring match.
        # Substring matching is hard in plain Excel without arrays; fallback: explicit IF chain on the 5 known categories.
        f_prog_simple = (
            f'=IF(ISNUMBER(SEARCH("232",{fha_col}{r})),VLOOKUP("232",V7_PROGRAM_TABLE,2,FALSE),'
            f'IF(ISNUMBER(SEARCH("538",{fha_col}{r})),VLOOKUP("538",V7_PROGRAM_TABLE,2,FALSE),'
            f'IF(ISNUMBER(SEARCH("223a7",{fha_col}{r})),VLOOKUP("223a7",V7_PROGRAM_TABLE,2,FALSE),'
            f'IF(ISNUMBER(SEARCH("223f",{fha_col}{r})),VLOOKUP("223f",V7_PROGRAM_TABLE,2,FALSE),'
            f'V7_PROGRAM_DEFAULT))))'
        )
        ws.cell(r, M['M_program'], f_prog_simple).number_format = "0.0000"

        # M_purpose: triangular hump centered at peak_age, half-width `width`
        is_nc_col = cl(IN['is_NC_in'])
        f_purp = (f"=IF({is_nc_col}{r}=1,"
                  f"1+V7_PURPOSE_NC_BUMP*MAX(0,1-ABS({age_col}{r}-V7_PURPOSE_PEAK_AGE)/V7_PURPOSE_WIDTH),"
                  f"1)")
        ws.cell(r, M['M_purpose'], f_purp).number_format = "0.0000"

        # M_lockout: msle from period and lockout_end YYYYMMs; spike-and-decay
        period_col = cl(ID['period'])
        lo_col = cl(IN['lockout_end_yyyymm'])
        # Compute msle inline; if lockout_end is blank/non-numeric, msle = 0
        msle_expr = (
            f'IF(OR({lo_col}{r}="",NOT(ISNUMBER({lo_col}{r}))),0,'
            f'MAX(0,MIN(24,'
            f'(INT(VALUE({period_col}{r})/100)-INT({lo_col}{r}/100))*12'
            f'+MOD(VALUE({period_col}{r}),100)-MOD({lo_col}{r},100))))'
        )
        f_lock = (f"=IF(AND({msle_expr}>0,{msle_expr}<=12),"
                  f"1+V7_LOCKOUT_AMP*EXP(-({msle_expr})/V7_LOCKOUT_TAU),1)")
        ws.cell(r, M['M_lockout'], f_lock).number_format = "0.0000"

        # M_maturity: ramp; mtm computed from maturity YYYYMM and period
        mat_col = cl(IN['loan_maturity_yyyymm'])
        mtm_expr = (
            f'IF(OR({mat_col}{r}="",NOT(ISNUMBER({mat_col}{r}))),360,'
            f'MAX(0,MIN(480,'
            f'(INT({mat_col}{r}/100)-INT(VALUE({period_col}{r})/100))*12'
            f'+MOD({mat_col}{r},100)-MOD(VALUE({period_col}{r}),100))))'
        )
        f_mty = (f"=IF({mtm_expr}>0,"
                 f"1+V7_MTY_AMP*MAX(0,(V7_MTY_CUTOFF-({mtm_expr}))/V7_MTY_CUTOFF),"
                 f"1)")
        ws.cell(r, M['M_maturity'], f_mty).number_format = "0.0000"

        # M_burnout: max(floor, 1 - slope*burn_ratio)
        br_col = cl(IN['burn_ratio_in'])
        f_burn = (f"=MAX(V7_BURN_FLOOR,1-V7_BURN_SLOPE*MIN(1,MAX(0,{br_col}{r})))")
        ws.cell(r, M['M_burnout'], f_burn).number_format = "0.0000"

        # M_issuer overlay (default 1.0 unless toggle is on)
        iss_col = cl(IN['issuer_number'])
        f_iss = (
            f'=IF(V7_USE_ISSUER=1,IFERROR(VLOOKUP({iss_col}{r},V7_ISSUER_TABLE,3,FALSE),1),1)'
        )
        ws.cell(r, M['M_issuer'], f_iss).number_format = "0.0000"

        # ----- Summary block -----
        # product = M_age × M_rate × M_size × M_program × M_purpose × M_lockout × M_maturity × M_burnout × M_issuer
        first_m = cl(M['M_age'])
        last_m  = cl(M['M_issuer'])
        ws.cell(r, SU['product'],
                f"=PRODUCT({first_m}{r}:{last_m}{r})").number_format = "0.0000"
        ws.cell(r, SU['capped_SMM'],
                f"=MIN(V7_BASE_SMM*{cl(SU['product'])}{r},V7_SMM_CAP)").number_format = "0.000000"
        ws.cell(r, SU['pred_CPR'],
                f"=1-(1-{cl(SU['capped_SMM'])}{r})^12").number_format = "0.00%"

        # Attribution: log-multiplicative decomposition (uses CORE 9 multipliers, not the issuer overlay)
        # log_mult_j = LN(M_j); attribution_j = log_mult_j / SUM(log_mults_core) × (pred_CPR - base_CPR)
        log_terms = "+".join(f"LN({cl(M[name])}{r})" for name in v7m.MULTIPLIER_NAMES)
        sum_log_expr = f"({log_terms})"
        for name in v7m.MULTIPLIER_NAMES:
            attr_col_idx = AT[f"attr_{name}"]
            f_attr = (
                f"=IFERROR(LN({cl(M[name])}{r})/{sum_log_expr}"
                f"*({cl(SU['pred_CPR'])}{r}-V7_BASE_CPR),0)"
            )
            ws.cell(r, attr_col_idx, f_attr).number_format = "0.00%"

        # Sanity: base_CPR + SUM(attr) - pred_CPR
        first_a = cl(AT[f"attr_{v7m.MULTIPLIER_NAMES[0]}"])
        last_a  = cl(AT[f"attr_{v7m.MULTIPLIER_NAMES[-1]}"])
        ws.cell(r, COL_SANITY,
                f"=V7_BASE_CPR+SUM({first_a}{r}:{last_a}{r})-{cl(SU['pred_CPR'])}{r}"
                ).number_format = "0.000000%"

    # Column widths
    for j in range(1, COL_SANITY + 1):
        if j <= n_id:
            ws.column_dimensions[cl(j)].width = 22 if j == ID['loan_id'] else 16
        elif j < COL_DERIVED_START:
            ws.column_dimensions[cl(j)].width = 14
        elif j < COL_MULT_START:
            ws.column_dimensions[cl(j)].width = 12
        elif j < COL_SUMM_START:
            ws.column_dimensions[cl(j)].width = 11
        elif j < COL_ATTR_START:
            ws.column_dimensions[cl(j)].width = 12
        else:
            ws.column_dimensions[cl(j)].width = 14
    ws.freeze_panes = ws.cell(DATA_START_ROW, COL_MULT_START).coordinate


# ---------------------------------------------------------------------------
# Sheet 5: Deal_Aggregator
# ---------------------------------------------------------------------------
def build_deal_aggregator(wb, sample, params):
    ws = wb.create_sheet("Deal_Aggregator")
    ws.cell(1, 1, "Deal-level Rollup (UPB-weighted)").font = Font(bold=True, size=14)
    ws.cell(2, 1, "Deal_ID is sourced from Loan_Snapshot column B. Edit there to regroup.").font = Font(italic=True, color="6B7280")

    # Compute Loan_Snapshot column letters
    ID_END = 8
    IN_START = 9
    DR_START = IN_START + 11   # 20
    M_START  = DR_START + 2    # 22
    SU_START = M_START + 10    # 32
    AT_START = SU_START + 3    # 35
    SAN = AT_START + 9         # 44
    UPB_COL = cl(IN_START + 0)
    DEAL_COL = "B"
    CPR_COL = cl(SU_START + 2)

    LS = "Loan_Snapshot"
    DATA_START_ROW = 3
    N = len(sample)
    LS_DEAL = f"{LS}!${DEAL_COL}${DATA_START_ROW}:${DEAL_COL}${DATA_START_ROW + N - 1}"
    LS_UPB  = f"{LS}!${UPB_COL}${DATA_START_ROW}:${UPB_COL}${DATA_START_ROW + N - 1}"
    LS_CPR  = f"{LS}!${CPR_COL}${DATA_START_ROW}:${CPR_COL}${DATA_START_ROW + N - 1}"

    # Headers
    hdrs = ["Deal_ID", "n_loans", "total_UPB", "wavg_pred_CPR"]
    hdrs += [f"attr_{n}" for n in v7m.MULTIPLIER_NAMES]
    hdrs += ["sanity"]
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(4, j, h); c.fill = HEADER_FILL; c.font = WHITE; c.alignment = CENTER

    # Default deal grouping
    sd = sample.copy()
    sd['_iss'] = sd['issuer_name'].fillna('').astype(str).str.split().str[0].str.replace(',', '').str.upper().str[:12]
    sd['_era'] = np.where(sd['vintage_year'] >= 2021, 'POST21', 'PRE21')
    sd['_deal'] = sd['_iss'] + '_' + sd['_era']
    deals = sd['_deal'].value_counts().index.tolist()

    agg_start = 5
    for i, deal in enumerate(deals):
        r = agg_start + i
        ws.cell(r, 1, str(deal)).font = MONO
        ws.cell(r, 2, f'=COUNTIF({LS_DEAL},$A{r})')
        ws.cell(r, 3, f'=SUMIF({LS_DEAL},$A{r},{LS_UPB})').number_format = "#,##0"
        ws.cell(r, 4, f'=IFERROR(SUMPRODUCT(({LS_DEAL}=$A{r})*{LS_UPB}*{LS_CPR})/$C{r},0)').number_format = "0.00%"
        for k, name in enumerate(v7m.MULTIPLIER_NAMES):
            attr_letter = cl(AT_START + k)
            LS_ATTR = f"{LS}!${attr_letter}${DATA_START_ROW}:${attr_letter}${DATA_START_ROW + N - 1}"
            ws.cell(r, 5 + k, f'=IFERROR(SUMPRODUCT(({LS_DEAL}=$A{r})*{LS_UPB}*{LS_ATTR})/$C{r},0)').number_format = "0.00%"
        first_a = cl(5)
        last_a  = cl(5 + len(v7m.MULTIPLIER_NAMES) - 1)
        ws.cell(r, 5 + len(v7m.MULTIPLIER_NAMES),
                f"=V7_BASE_CPR+SUM({first_a}{r}:{last_a}{r})-$D{r}").number_format = "0.000000%"

    # ALL row
    total_row = agg_start + len(deals) + 2
    ws.cell(total_row, 1, "ALL LOANS (snapshot total)").font = BOLD
    ws.cell(total_row, 2, f'=COUNTA({LS}!$A${DATA_START_ROW}:$A${DATA_START_ROW + N - 1})')
    ws.cell(total_row, 3, f'=SUM({LS_UPB})').number_format = "#,##0"
    ws.cell(total_row, 4, f'=SUMPRODUCT({LS_UPB}*{LS_CPR})/$C{total_row}').number_format = "0.00%"
    for k, name in enumerate(v7m.MULTIPLIER_NAMES):
        attr_letter = cl(AT_START + k)
        LS_ATTR = f"{LS}!${attr_letter}${DATA_START_ROW}:${attr_letter}${DATA_START_ROW + N - 1}"
        ws.cell(total_row, 5 + k, f'=SUMPRODUCT({LS_UPB}*{LS_ATTR})/$C{total_row}').number_format = "0.00%"
    first_a = cl(5)
    last_a  = cl(5 + len(v7m.MULTIPLIER_NAMES) - 1)
    ws.cell(total_row, 5 + len(v7m.MULTIPLIER_NAMES),
            f"=V7_BASE_CPR+SUM({first_a}{total_row}:{last_a}{total_row})-$D{total_row}").number_format = "0.000000%"
    for j in range(1, 5 + len(v7m.MULTIPLIER_NAMES) + 1):
        ws.cell(total_row, j).fill = BASELINE_FILL

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    for k in range(len(v7m.MULTIPLIER_NAMES) + 1):
        ws.column_dimensions[cl(5 + k)].width = 13
    ws.freeze_panes = "B5"


# ---------------------------------------------------------------------------
# Sheet 6: Benchmarks (SanCap reference grid)
# ---------------------------------------------------------------------------
def build_benchmarks_tab(wb, params, md):
    ws = wb.create_sheet("Benchmarks")
    ws.cell(1, 1, "SanCap (Amherst Pierpont 2014-2018) reference benchmarks vs V7").font = Font(bold=True, size=14)
    ws.cell(2, 1, "Informational sanity check — V7 trains on the 2018-2026 panel which spans the 2022-2024 rate-hike cycle.").font = Font(italic=True, color="6B7280")
    ws.cell(3, 1, "Persistent divergences in +50/+100bp scenarios reflect that regime shift, not a model defect. Use V7 as ground truth on this panel; SanCap is a shape-direction sanity check.").font = Font(italic=True, color="6B7280")

    grid = md.get('sancap_benchmarks', [])
    hdrs = ["Scenario", "SanCap CPR (2014-2018)", "V7 pred CPR", "Δ (pp)", "Flag"]
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(4, j, h); c.fill = HEADER_FILL; c.font = WHITE; c.alignment = CENTER

    for i, row in enumerate(grid):
        r = 5 + i
        ws.cell(r, 1, row['scenario'])
        ws.cell(r, 2, row['sancap_cpr']).number_format = "0.0"
        ws.cell(r, 3, row['v7_pred_cpr']).number_format = "0.00"
        ws.cell(r, 4, row['divergence_pp']).number_format = "0.00"
        flag_cell = ws.cell(r, 5, row['flag'])
        if row['flag'] == 'REVIEW':
            flag_cell.fill = PatternFill("solid", fgColor="FECACA")
            flag_cell.font = Font(bold=True, color="991B1B")
        else:
            flag_cell.fill = PatternFill("solid", fgColor="D1FAE5")

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12


if __name__ == '__main__':
    main()
