"""
Build the V6E or V6F Excel calculator workbook.

Selects model via env var MODEL_VERSION (default v6e). Both produce the same
4-tab layout (README / Coefficients / Loan_Snapshot / Deal_Aggregator); only
the feature list, derivations, and per-loan input columns differ.

Math form (identical for both): centered native units.
    contribution_j = beta_native_j * (x_j - mean_j)
    total_logit_dev = SUM(contributions)
    logit_z = intercept_scaled + total_logit_dev
    SMM = 1 / (1 + EXP(-logit_z))
    pred_CPR = 1 - (1-SMM)^12
    baseline_CPR = 1 - (1 - SIGMOID(intercept_scaled))^12   # CPR at feature means
    attribution_j = (contribution_j / total_logit_dev) * (pred_CPR - baseline_CPR)

Usage:
    MODEL_VERSION=v6e python3 build_excel.py        # default
    MODEL_VERSION=v6f python3 build_excel.py
"""
import json, math, os, sys
from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

HERE = Path(__file__).resolve().parent
MODEL_VERSION = os.environ.get('MODEL_VERSION', 'v6e').lower()
if MODEL_VERSION not in ('v6e', 'v6f', 'v7', 'v8'):
    sys.exit(f"ERROR: unsupported MODEL_VERSION={MODEL_VERSION} (expected v6e, v6f, v7, or v8)")

# V7 and V8 use fundamentally different architectures (multiplicative vs cohort
# S-curve); each has its own dedicated builder.
if MODEL_VERSION == 'v7':
    import build_excel_v7
    build_excel_v7.main()
    sys.exit(0)
if MODEL_VERSION == 'v8':
    import build_excel_v8
    build_excel_v8.main()
    sys.exit(0)

MODEL_JSON = HERE / f"model_data_{MODEL_VERSION}.json"
SAMPLE_PARQUET = HERE / "sample_loans.parquet"
OUT_XLSX = HERE / f"{MODEL_VERSION.upper()}_Excel_Calculator.xlsx"

if not MODEL_JSON.exists():
    sys.exit(f"ERROR: {MODEL_JSON} not found - run train_{MODEL_VERSION}.py first")

# ---------- Load model ----------
md = json.load(open(MODEL_JSON))
coefs = {c['feature']: c for c in md['coefficients']}
intercept_scaled = md['metadata']['intercept_scaled']
feature_list = md['metadata']['feature_list']  # canonical order

print(f"Building {OUT_XLSX.name} from {MODEL_JSON.name}")
print(f"  Model: {MODEL_VERSION.upper()}  features={len(feature_list)}  test_AUC={md['metadata']['test_auc']:.4f}")


def regenerate_sample(panel_path):
    """Pull a stratified ~250-loan sample from the latest non-excluded period of a panel."""
    print(f"  Reading panel: {panel_path}")
    df = pd.read_parquet(panel_path)
    df = df.sort_values(['loan_id', 'period'])
    df['itm'] = (df['refi_incentive_bps'] > 0).astype(int)
    df['cum_itm'] = df.groupby('loan_id')['itm'].cumsum()
    # Snapshot period = last period in panel minus 1 (last is excluded by Fix 9 gating)
    snap_period = sorted(df['period'].astype(str).unique())[-2]
    print(f"  Snapshot period: {snap_period}")
    snap = df[(df['period'].astype(str) == snap_period)
              & (df['prepay_eligible'] == 1) & (df['has_valid_data'] == 1)].copy()
    buckets = []
    for issuer_pat, n in [
        ('LUMENT', 35), ('BERKADIA', 30), ('WALKER', 25), ('GREYSTONE', 25),
        ('MERCHANTS', 20), ('PNC', 20), ('WELLS FARGO', 20), ('CAPITAL FUNDING', 20),
        ('DWIGHT', 15), ('KEYBANK', 15), ('JLL', 10), ('BELLWETHER', 8),
        ('PGIM', 5), ('NORTHMARQ', 5),
    ]:
        m = snap['issuer_name'].fillna('').str.upper().str.contains(issuer_pat, na=False, regex=False)
        sub = snap[m]
        if len(sub) > 0:
            buckets.append(sub.sample(min(n, len(sub)), random_state=42))
    out = pd.concat(buckets).drop_duplicates('loan_id').reset_index(drop=True)
    out.to_parquet(SAMPLE_PARQUET)
    print(f"  Wrote {SAMPLE_PARQUET.name} (n={len(out)})")
    return out


# ---------- Resolve sample ----------
panel_env = os.environ.get('GNMA_PANEL_PARQUET')
if panel_env and Path(panel_env).exists():
    sample = regenerate_sample(panel_env).reset_index(drop=True)
elif SAMPLE_PARQUET.exists():
    sample = pd.read_parquet(SAMPLE_PARQUET).reset_index(drop=True)
else:
    sys.exit(f"ERROR: neither GNMA_PANEL_PARQUET set nor {SAMPLE_PARQUET} present.")

# Drop rows with NaN loan_age_months or vintage_year (rare panel-corruption artifacts).
# Without this filter, Excel renders the NaN as an empty cell which then
# produces a degenerate burn_ratio. Documented data-quality quirk: ~2 loans.
before = len(sample)
sample = sample[sample['loan_age_months'].notna() & sample['vintage_year'].notna()].reset_index(drop=True)
if len(sample) < before:
    print(f"  Dropped {before - len(sample)} loans with NaN age/vintage; remaining: {len(sample)}")

# ---------- Style helpers ----------
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
SUBHEADER_FILL = PatternFill("solid", fgColor="374151")
PARAM_FILL = PatternFill("solid", fgColor="FEF3C7")
FORMULA_FILL = PatternFill("solid", fgColor="E0F2FE")
BASELINE_FILL = PatternFill("solid", fgColor="FFEDD5")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
MONO = Font(name="Consolas")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

thin = Side(border_style="thin", color="9CA3AF")
BORDER = Border(top=thin, bottom=thin, left=thin, right=thin)

wb = Workbook()
wb.remove(wb.active)

# ============================================================
# Sheet 1: README
# ============================================================
ws = wb.create_sheet("README")
readme_rows = [
    (f"{MODEL_VERSION.upper()} Voluntary Prepayment Calculator", BOLD),
    ("", None),
    (f"Model: {MODEL_VERSION.upper()} ({len(feature_list)}-feature logistic regression, test AUC {md['metadata']['test_auc']:.4f})", None),
    (f"Training period: {md['metadata']['period_range'][0]} to {md['metadata']['period_range'][1]}", None),
    (f"Training events: {md['metadata']['training_events']:,} voluntary prepays in {md['metadata']['training_pop_n']:,} eligible loan-months", None),
    ("", None),
    ("Tabs", BOLD),
    (f"  Coefficients     - {len(feature_list)} feature parameters + intercept + diagnostics. Edit at your own risk.", None),
    ("  Loan_Snapshot    - one row per loan, current-period CPR + per-feature attribution.", None),
    ("                     Paste new loans into rows below the existing data; formulas autopopulate.", None),
    ("  Deal_Aggregator  - UPB-weighted deal-level rollup. Edit Deal_ID column in Loan_Snapshot to control grouping.", None),
    ("", None),
    ("Math (centered native form)", BOLD),
    ("  contribution_j = beta_native_j * (x_j - mean_j)", MONO),
    ("  logit_z        = intercept_scaled + SUM(contributions)", MONO),
    ("  SMM            = 1 / (1 + EXP(-logit_z))", MONO),
    ("  pred_CPR       = 1 - (1 - SMM)^12", MONO),
    ("  baseline_CPR   = CPR of a hypothetical loan with all features at their training-mean values", MONO),
    ("  attribution_j  = (contribution_j / SUM(contributions)) * (pred_CPR - baseline_CPR)", MONO),
    ("", None),
    ("Sanity check (column 'Sanity' in Loan_Snapshot)", BOLD),
    ("  baseline_CPR + SUM(attribution) - pred_CPR  should equal 0 (within rounding).", MONO),
    ("", None),
    ("Scope and caveats", BOLD),
    ("  - Snapshot CPR only. Forward projection (rate-shock vectors, monthly amortization) is NOT in this", None),
    ("    workbook; it lives in the dashboard's Loan Projector tab.", None),
    ("  - Refi incentive is RECOMPUTED in Excel from loan_rate, plc_rate, and prepay_penalty_points so", None),
    ("    you can run what-ifs by editing those inputs.", None),
    ("  - cum_itm is a STATE variable carried in from the panel. To do forward what-ifs, you must roll it", None),
    ("    forward yourself (add 1 each month the projected refi_incentive > 0).", None),
    ("  - is_post_covid is currently a vintage>=2021 flag. It captures real cohort behavior today but is", None),
    ("    fragile in low-rate forward scenarios.", None),
    ("  - Issuer dummies cover 7 names; all other issuers (incl. Lument, Greystone) sit in 'other' baseline.", None),
    ("", None),
    ("File built from", BOLD),
    (f"  Panel: gnma_mf_raw_data_20260504_135955.parquet (1,314,304 rows, 88 cols)", MONO),
    (f"  Model: model_data_v6e.json (snapshot {md['metadata']['last_period']})", MONO),
]
for i, (text, font) in enumerate(readme_rows, 1):
    cell = ws.cell(row=i, column=1, value=text)
    if font is not None:
        cell.font = font
ws.column_dimensions['A'].width = 110

# ============================================================
# Sheet 2: Coefficients
# ============================================================
ws = wb.create_sheet("Coefficients")
ws.cell(1, 1, f"{MODEL_VERSION.upper()} Voluntary Prepayment Model - Coefficients").font = Font(bold=True, size=14)
ws.cell(2, 1, "Edit at your own risk. Loan_Snapshot pulls mean/std/beta_native from this tab.").font = Font(italic=True, color="6B7280")

headers = ["#", "feature", "group", "beta_scaled", "beta_native", "mean", "std", "label"]
for j, h in enumerate(headers, 1):
    c = ws.cell(4, j, h); c.fill = HEADER_FILL; c.font = WHITE; c.alignment = CENTER

groups = md['metadata']['feature_groups']
labels = md['metadata']['feature_labels']
def feature_group(f):
    for g, items in groups.items():
        if f in items: return g
    return ""

# Sort by group then by feature_list order
for i, f in enumerate(feature_list, 1):
    c = coefs[f]
    row = 4 + i
    ws.cell(row, 1, i)
    ws.cell(row, 2, f).font = MONO
    ws.cell(row, 3, feature_group(f))
    ws.cell(row, 4, round(c['beta_scaled'], 6))
    ws.cell(row, 5, round(c['beta_native'], 8))
    ws.cell(row, 6, round(c['mean'], 6))
    ws.cell(row, 7, round(c['std'], 6))
    ws.cell(row, 8, labels.get(f, ""))
    for col in range(1, 9):
        ws.cell(row, col).border = BORDER
        if col in (4,5,6,7):
            ws.cell(row, col).number_format = "0.000000"
            ws.cell(row, col).font = MONO

n_feat = len(feature_list)
last_feat_row = 4 + n_feat

# Constants block
const_start = last_feat_row + 3
ws.cell(const_start, 1, "Model constants").font = BOLD
ws.cell(const_start+1, 1, "intercept_scaled")
ws.cell(const_start+1, 2, round(intercept_scaled, 6)); ws.cell(const_start+1, 2).font = MONO
ws.cell(const_start+2, 1, "baseline_SMM (sigmoid of intercept_scaled)")
ws.cell(const_start+2, 2, f"=1/(1+EXP(-B{const_start+1}))"); ws.cell(const_start+2, 2).font = MONO; ws.cell(const_start+2, 2).number_format = "0.00000000"
ws.cell(const_start+3, 1, "baseline_CPR (annualized from baseline_SMM)")
ws.cell(const_start+3, 2, f"=1-(1-B{const_start+2})^12"); ws.cell(const_start+3, 2).font = MONO; ws.cell(const_start+3, 2).number_format = "0.0000%"
ws.cell(const_start+5, 1, "Diagnostics").font = BOLD
ws.cell(const_start+6, 1, "test_AUC"); ws.cell(const_start+6, 2, round(md['metadata']['test_auc'], 4))
ws.cell(const_start+7, 1, "panel_mean_CPR"); ws.cell(const_start+7, 2, round(md['metadata']['base_cpr']/100, 6)); ws.cell(const_start+7, 2).number_format = "0.0000%"
ws.cell(const_start+8, 1, "training_pop_n"); ws.cell(const_start+8, 2, md['metadata']['training_pop_n'])
ws.cell(const_start+9, 1, "training_events"); ws.cell(const_start+9, 2, md['metadata']['training_events'])
ws.cell(const_start+10, 1, "panel_period_range"); ws.cell(const_start+10, 2, f"{md['metadata']['period_range'][0]} - {md['metadata']['period_range'][1]}")

for c in [const_start+1, const_start+2, const_start+3, const_start+6, const_start+7, const_start+8, const_start+9, const_start+10]:
    ws.cell(c, 1).font = Font(italic=True)
ws.cell(const_start+1, 2).fill = PARAM_FILL
ws.cell(const_start+2, 2).fill = BASELINE_FILL
ws.cell(const_start+3, 2).fill = BASELINE_FILL

# Define names for cross-sheet references
INTERCEPT_CELL = f"Coefficients!$B${const_start+1}"
BASELINE_SMM_CELL = f"Coefficients!$B${const_start+2}"
BASELINE_CPR_CELL = f"Coefficients!$B${const_start+3}"
wb.defined_names['INTERCEPT_SCALED'] = DefinedName('INTERCEPT_SCALED', attr_text=INTERCEPT_CELL)
wb.defined_names['BASELINE_SMM'] = DefinedName('BASELINE_SMM', attr_text=BASELINE_SMM_CELL)
wb.defined_names['BASELINE_CPR'] = DefinedName('BASELINE_CPR', attr_text=BASELINE_CPR_CELL)

# Column widths
widths = [4, 22, 26, 14, 16, 14, 14, 36]
for j, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(j)].width = w
ws.freeze_panes = "A5"

# Coefficient feature rows are at rows 5..(4+n_feat). Map feature -> row in Coefficients sheet.
COEF_ROW = {f: 4 + i + 1 for i, f in enumerate(feature_list)}  # row in 'Coefficients' sheet

# ============================================================
# Sheet 3: Loan_Snapshot
# ============================================================
ws = wb.create_sheet("Loan_Snapshot")

# Layout columns:
ID_COLS = ["loan_id", "deal_id", "pool_cusip", "issuer_name", "fha_category", "loan_purpose", "vintage_year", "period"]
if MODEL_VERSION == 'v6e':
    INPUT_COLS = ["upb", "loan_rate_pct", "plc_rate_bps", "prepay_penalty_points", "loan_age_months_in", "cum_itm_in", "in_prepay_penalty_in",
                  "is_hc_232_raw", "is_223a7_raw", "is_538_raw", "is_new_construction_raw"]
else:  # v6f
    # V6F adds: loan_maturity_date_int (YYYYMM int form), lockout_end_date_int, issuer_number_str.
    # Drops cum_itm_in and in_prepay_penalty_in (no longer used as features), but keeps cum_itm_in
    # because burn_ratio still needs it as an input state variable.
    INPUT_COLS = ["upb", "loan_rate_pct", "plc_rate_bps", "prepay_penalty_points", "loan_age_months_in",
                  "cum_itm_in", "loan_maturity_yyyymm", "lockout_end_yyyymm", "issuer_number_str",
                  "is_hc_232_raw", "is_223a7_raw", "is_538_raw", "is_new_construction_raw"]
# Feature columns (canonical order): each becomes a column. Some are formulas referencing inputs.

n_id = len(ID_COLS)
n_input = len(INPUT_COLS)

# Column positions
COL_ID_START = 1
COL_INPUT_START = COL_ID_START + n_id              # 1..8 IDs
COL_FEAT_START = COL_INPUT_START + n_input         # features start here
COL_CONTRIB_START = COL_FEAT_START + n_feat        # 18 contribution cols
COL_SUMMARY_START = COL_CONTRIB_START + n_feat     # total_logit_dev, logit_z, SMM, pred_CPR
COL_ATTR_START = COL_SUMMARY_START + 4
COL_SANITY = COL_ATTR_START + n_feat

# Row layout:
# Row 1: section banners
# Row 2: column headers (feature name, etc.)
# Row 3: parameter row 1 - mean (for feature columns); blank elsewhere
# Row 4: parameter row 2 - beta_native (for feature columns); blank elsewhere
# Row 5+: data

# Banners
def banner(col_start, col_end, text, fill=SUBHEADER_FILL):
    ws.merge_cells(start_row=1, end_row=1, start_column=col_start, end_column=col_end)
    c = ws.cell(1, col_start, text); c.fill = fill; c.font = WHITE; c.alignment = CENTER

banner(COL_ID_START, COL_INPUT_START - 1, "Loan Identifiers")
banner(COL_INPUT_START, COL_FEAT_START - 1, "Raw Inputs (editable)", PatternFill("solid", fgColor="064E3B"))
banner(COL_FEAT_START, COL_CONTRIB_START - 1, "V6E Feature Values (formulas - do not edit)")
banner(COL_CONTRIB_START, COL_SUMMARY_START - 1, "Logit Contributions (β_native × (x − mean))", PatternFill("solid", fgColor="312E81"))
banner(COL_SUMMARY_START, COL_ATTR_START - 1, "CPR Summary", PatternFill("solid", fgColor="92400E"))
banner(COL_ATTR_START, COL_SANITY - 1, "CPR Attribution (proportional method)", PatternFill("solid", fgColor="9D174D"))
banner(COL_SANITY, COL_SANITY, "Check", PatternFill("solid", fgColor="6B7280"))

# Headers row 2
hdrs = []
hdrs += ID_COLS
hdrs += INPUT_COLS
hdrs += feature_list  # 18 V6E features
hdrs += [f"contrib_{f}" for f in feature_list]
hdrs += ["total_logit_dev", "logit_z", "SMM", "pred_CPR"]
hdrs += [f"attr_{f}" for f in feature_list]
hdrs += ["Sanity"]
for j, h in enumerate(hdrs, 1):
    c = ws.cell(2, j, h)
    c.fill = HEADER_FILL; c.font = WHITE; c.alignment = CENTER

# Parameter rows 3 (mean) and 4 (beta_native): only for feature columns
for k, f in enumerate(feature_list):
    col = COL_FEAT_START + k
    ws.cell(3, col, f"=Coefficients!$F${COEF_ROW[f]}").number_format = "0.000000"
    ws.cell(4, col, f"=Coefficients!$E${COEF_ROW[f]}").number_format = "0.00000000"
    ws.cell(3, col).fill = PARAM_FILL; ws.cell(4, col).fill = PARAM_FILL
    ws.cell(3, col).font = MONO; ws.cell(4, col).font = MONO
ws.cell(3, 1, "mean").font = BOLD; ws.cell(4, 1, "β_native").font = BOLD

# Helper: column letters
def cl(idx): return get_column_letter(idx)
FEAT_COL = {f: COL_FEAT_START + k for k, f in enumerate(feature_list)}
CONTRIB_COL = {f: COL_CONTRIB_START + k for k, f in enumerate(feature_list)}
ATTR_COL = {f: COL_ATTR_START + k for k, f in enumerate(feature_list)}

ID_LOAN, ID_DEAL, ID_CUSIP, ID_ISSUER, ID_FHA, ID_LP, ID_VY, ID_PERIOD = range(1, 9)
IN_UPB = COL_INPUT_START + 0
IN_LOANRATE = COL_INPUT_START + 1
IN_PLCBPS = COL_INPUT_START + 2
IN_PPP = COL_INPUT_START + 3
IN_AGE = COL_INPUT_START + 4
IN_CUMITM = COL_INPUT_START + 5
if MODEL_VERSION == 'v6e':
    IN_INPEN = COL_INPUT_START + 6
    IN_HC232 = COL_INPUT_START + 7
    IN_223A7 = COL_INPUT_START + 8
    IN_538   = COL_INPUT_START + 9
    IN_NC    = COL_INPUT_START + 10
else:  # v6f
    IN_MTRDATE  = COL_INPUT_START + 6   # YYYYMM int (e.g. 204012 = Dec 2040)
    IN_LOENDDATE = COL_INPUT_START + 7
    IN_ISSNUM   = COL_INPUT_START + 8
    IN_HC232    = COL_INPUT_START + 9
    IN_223A7    = COL_INPUT_START + 10
    IN_538      = COL_INPUT_START + 11
    IN_NC       = COL_INPUT_START + 12

DATA_START_ROW = 5
N_LOANS = len(sample)

# Issuer name patterns for dummy columns
if MODEL_VERSION == 'v6e':
    ISSUER_PATTERNS = {
        'iss_capital_funding': 'CAPITAL FUNDING',
        'iss_merchants':       'MERCHANTS',
        'iss_wells_fargo':     'WELLS FARGO',
        'iss_walker_dunlop':   'WALKER',
        'iss_pnc':             'PNC',
        'iss_berkadia':        'BERKADIA',
        'iss_dwight':          'DWIGHT',
    }
else:  # v6f
    ISSUER_PATTERNS = {
        'iss_capital_funding': 'CAPITAL FUNDING',
        'iss_wells_fargo':     'WELLS FARGO',
        'iss_pnc':             'PNC',
        'iss_dwight':          'DWIGHT',
        'iss_greystone':       'GREYSTONE',
        # iss_lument_combined uses issuer_number, handled below
    }

def parse_yyyymmdd_to_yyyymm(date_str):
    """Convert YYYYMMDD string to YYYYMM int. Returns None if unparseable."""
    if not date_str or pd.isna(date_str):
        return None
    s = str(date_str).strip()
    if len(s) >= 6 and s[:4].isdigit() and s[4:6].isdigit():
        return int(s[:6])
    return None


for i in range(N_LOANS):
    r = DATA_START_ROW + i
    s = sample.iloc[i]
    # Identifiers
    ws.cell(r, ID_LOAN, str(s['loan_id'])).font = MONO
    # Default Deal_ID groups by issuer + vintage era for a useful demo aggregation.
    issuer_short = str(s['issuer_name']).split()[0].replace(',', '').upper()[:12]
    era = "POST21" if (pd.notna(s['vintage_year']) and s['vintage_year'] >= 2021) else "PRE21"
    deal_default = f"{issuer_short}_{era}"
    ws.cell(r, ID_DEAL, deal_default)
    ws.cell(r, ID_CUSIP, str(s['pool_cusip'])).font = MONO
    ws.cell(r, ID_ISSUER, str(s['issuer_name']))
    ws.cell(r, ID_FHA, str(s.get('fha_category', '')))
    ws.cell(r, ID_LP, str(s.get('loan_purpose', '')))
    ws.cell(r, ID_VY, int(s['vintage_year']) if pd.notna(s['vintage_year']) else None)
    ws.cell(r, ID_PERIOD, str(s['period']))
    # Raw inputs (editable)
    ws.cell(r, IN_UPB, float(s['upb'])).number_format = "#,##0"
    ws.cell(r, IN_LOANRATE, float(s['loan_rate'])).number_format = "0.000"
    ws.cell(r, IN_PLCBPS, float(s['plc_rate_bps'])).number_format = "0.0"
    ws.cell(r, IN_PPP, float(s.get('prepay_penalty_points', 0)) if pd.notna(s.get('prepay_penalty_points', 0)) else 0).number_format = "0.0"
    ws.cell(r, IN_AGE, float(s['loan_age_months']))
    ws.cell(r, IN_CUMITM, int(s['cum_itm']))
    ws.cell(r, IN_HC232, int(s.get('is_hc_232', 0)))
    ws.cell(r, IN_223A7, int(s.get('is_223a7', 0)))
    ws.cell(r, IN_538, int(s.get('is_538', 0)))
    ws.cell(r, IN_NC, int(s.get('is_new_construction', 0)))

    if MODEL_VERSION == 'v6e':
        # V6E-only inputs
        ws.cell(r, IN_INPEN, int(s['in_prepay_penalty']))
        # V6E feature formulas
        # 1. refi_incentive_bps = loan_rate*100 - (plc_bps + (1+ppp)*12.5)
        ws.cell(r, FEAT_COL['refi_incentive_bps'],
                f"={cl(IN_LOANRATE)}{r}*100-({cl(IN_PLCBPS)}{r}+(1+{cl(IN_PPP)}{r})*12.5)").number_format = "0.0"
        # 2. in_prepay_penalty (passthrough)
        ws.cell(r, FEAT_COL['in_prepay_penalty'], f"={cl(IN_INPEN)}{r}")
        # 3. loan_age_months
        ws.cell(r, FEAT_COL['loan_age_months'], f"={cl(IN_AGE)}{r}")
        # 4. log_upb
        ws.cell(r, FEAT_COL['log_upb'], f"=LN(1+{cl(IN_UPB)}{r})").number_format = "0.000"
        # 5. is_post_covid
        ws.cell(r, FEAT_COL['is_post_covid'], f"=IF({cl(ID_VY)}{r}>=2021,1,0)")
        # 6. cum_itm
        ws.cell(r, FEAT_COL['cum_itm'], f"={cl(IN_CUMITM)}{r}")
        # 7. burn_ratio
        ws.cell(r, FEAT_COL['burn_ratio'],
            f'=IF(ISNUMBER({cl(IN_AGE)}{r}),{cl(IN_CUMITM)}{r}/MAX({cl(IN_AGE)}{r},1),0)'
        ).number_format = "0.000"
        # 8-11. FHA/program flags
        ws.cell(r, FEAT_COL['is_hc_232'], f"={cl(IN_HC232)}{r}")
        ws.cell(r, FEAT_COL['is_223a7'], f"={cl(IN_223A7)}{r}")
        ws.cell(r, FEAT_COL['is_538'], f"={cl(IN_538)}{r}")
        ws.cell(r, FEAT_COL['lp_NC'], f"={cl(IN_NC)}{r}")
        # 12-18. Issuer dummies (V6E set)
        for fname, pat in ISSUER_PATTERNS.items():
            col = FEAT_COL[fname]
            ws.cell(r, col, f'=IF(ISNUMBER(SEARCH("{pat}",{cl(ID_ISSUER)}{r})),1,0)')

    else:  # v6f
        # V6F-only inputs
        mtr_yyyymm = parse_yyyymmdd_to_yyyymm(s.get('loan_maturity_date', ''))
        lo_yyyymm  = parse_yyyymmdd_to_yyyymm(s.get('lockout_end_date', ''))
        ws.cell(r, IN_MTRDATE, mtr_yyyymm if mtr_yyyymm is not None else '')
        ws.cell(r, IN_LOENDDATE, lo_yyyymm if lo_yyyymm is not None else '')
        ws.cell(r, IN_ISSNUM, str(s.get('issuer_number', '')))
        # V6F feature formulas
        # 1. gross_refi_incentive_bps = loan_rate*100 - plc_bps   (penalty-neutral)
        ws.cell(r, FEAT_COL['gross_refi_incentive_bps'],
                f"={cl(IN_LOANRATE)}{r}*100-{cl(IN_PLCBPS)}{r}").number_format = "0.0"
        # 2. prepay_penalty_points (passthrough)
        ws.cell(r, FEAT_COL['prepay_penalty_points'], f"={cl(IN_PPP)}{r}").number_format = "0.0"
        # 3-5. Piecewise age
        ws.cell(r, FEAT_COL['age_0_36'],
                f"=MIN({cl(IN_AGE)}{r},36)").number_format = "0"
        ws.cell(r, FEAT_COL['age_36_120'],
                f"=MIN(MAX({cl(IN_AGE)}{r}-36,0),84)").number_format = "0"
        ws.cell(r, FEAT_COL['age_120plus'],
                f"=MAX({cl(IN_AGE)}{r}-120,0)").number_format = "0"
        # 6. months_to_maturity = (maturity_yyyymm year)*12 + month - (period year)*12 - month, capped 0..480
        # Period and maturity are YYYYMM integers like 202602 (year=2026, month=02)
        # If maturity input is blank, default to 360 (long-tail).
        mtm_formula = (
            f'=IF(OR({cl(IN_MTRDATE)}{r}="",NOT(ISNUMBER({cl(IN_MTRDATE)}{r}))),360,'
            f'MAX(0,MIN(480,'
            f'(INT({cl(IN_MTRDATE)}{r}/100)-INT(VALUE({cl(ID_PERIOD)}{r})/100))*12'
            f'+MOD({cl(IN_MTRDATE)}{r},100)-MOD(VALUE({cl(ID_PERIOD)}{r}),100))))'
        )
        ws.cell(r, FEAT_COL['months_to_maturity'], mtm_formula).number_format = "0"
        # 7. pre_maturity_flag = IF(0 < mtm <= 24, 1, 0)
        mtm_col = cl(FEAT_COL['months_to_maturity'])
        ws.cell(r, FEAT_COL['pre_maturity_flag'],
                f"=IF(AND({mtm_col}{r}>0,{mtm_col}{r}<=24),1,0)")
        # 8. months_since_lockout_end = MAX(0, MIN(24, period - lockout_end)); 0 if no date
        msle_formula = (
            f'=IF(OR({cl(IN_LOENDDATE)}{r}="",NOT(ISNUMBER({cl(IN_LOENDDATE)}{r}))),0,'
            f'MAX(0,MIN(24,'
            f'(INT(VALUE({cl(ID_PERIOD)}{r})/100)-INT({cl(IN_LOENDDATE)}{r}/100))*12'
            f'+MOD(VALUE({cl(ID_PERIOD)}{r}),100)-MOD({cl(IN_LOENDDATE)}{r},100))))'
        )
        ws.cell(r, FEAT_COL['months_since_lockout_end'], msle_formula).number_format = "0"
        # 9. log_upb
        ws.cell(r, FEAT_COL['log_upb'], f"=LN(1+{cl(IN_UPB)}{r})").number_format = "0.000"
        # 10-11. Size flags
        ws.cell(r, FEAT_COL['small_loan'], f"=IF({cl(IN_UPB)}{r}<2000000,1,0)")
        ws.cell(r, FEAT_COL['large_loan'], f"=IF({cl(IN_UPB)}{r}>50000000,1,0)")
        # 12. burn_ratio
        ws.cell(r, FEAT_COL['burn_ratio'],
            f'=IF(ISNUMBER({cl(IN_AGE)}{r}),{cl(IN_CUMITM)}{r}/MAX({cl(IN_AGE)}{r},1),0)'
        ).number_format = "0.000"
        # 13. is_post_covid
        ws.cell(r, FEAT_COL['is_post_covid'], f"=IF({cl(ID_VY)}{r}>=2021,1,0)")
        # 14-17. Program / purpose
        ws.cell(r, FEAT_COL['is_223a7'], f"={cl(IN_223A7)}{r}")
        ws.cell(r, FEAT_COL['is_538'], f"={cl(IN_538)}{r}")
        ws.cell(r, FEAT_COL['is_hc_232'], f"={cl(IN_HC232)}{r}")
        ws.cell(r, FEAT_COL['lp_NC'], f"={cl(IN_NC)}{r}")
        # 18-22. Issuer dummies (V6F set, by name)
        for fname, pat in ISSUER_PATTERNS.items():
            col = FEAT_COL[fname]
            ws.cell(r, col, f'=IF(ISNUMBER(SEARCH("{pat}",{cl(ID_ISSUER)}{r})),1,0)')
        # 23. iss_lument_combined: by issuer_number IN ('3896','3557')
        ws.cell(r, FEAT_COL['iss_lument_combined'],
                f'=IF(OR({cl(IN_ISSNUM)}{r}="3896",{cl(IN_ISSNUM)}{r}="3557"),1,0)')
        # 24+. Interactions (only emit those that are in feature_list)
        if 'gross_refi__x__prepay_pen' in FEAT_COL:
            ws.cell(r, FEAT_COL['gross_refi__x__prepay_pen'],
                    f"={cl(FEAT_COL['gross_refi_incentive_bps'])}{r}*{cl(FEAT_COL['prepay_penalty_points'])}{r}").number_format = "0.0"
        if 'age_0_36__x__lp_NC' in FEAT_COL:
            ws.cell(r, FEAT_COL['age_0_36__x__lp_NC'],
                    f"={cl(FEAT_COL['age_0_36'])}{r}*{cl(FEAT_COL['lp_NC'])}{r}").number_format = "0"
        if 'age_0_36__x__is_hc_232' in FEAT_COL:
            ws.cell(r, FEAT_COL['age_0_36__x__is_hc_232'],
                    f"={cl(FEAT_COL['age_0_36'])}{r}*{cl(FEAT_COL['is_hc_232'])}{r}").number_format = "0"
        if 'iss_lument__x__gross_refi' in FEAT_COL:
            ws.cell(r, FEAT_COL['iss_lument__x__gross_refi'],
                    f"={cl(FEAT_COL['iss_lument_combined'])}{r}*{cl(FEAT_COL['gross_refi_incentive_bps'])}{r}").number_format = "0.0"
        if 'iss_dwight__x__gross_refi' in FEAT_COL:
            ws.cell(r, FEAT_COL['iss_dwight__x__gross_refi'],
                    f"={cl(FEAT_COL['iss_dwight'])}{r}*{cl(FEAT_COL['gross_refi_incentive_bps'])}{r}").number_format = "0.0"
        if 'iss_greystone__x__gross_refi' in FEAT_COL:
            ws.cell(r, FEAT_COL['iss_greystone__x__gross_refi'],
                    f"={cl(FEAT_COL['iss_greystone'])}{r}*{cl(FEAT_COL['gross_refi_incentive_bps'])}{r}").number_format = "0.0"

    # Contribution columns: beta_native * (feature_value - mean)
    for k, f in enumerate(feature_list):
        fcol = FEAT_COL[f]
        ccol = CONTRIB_COL[f]
        # beta_native is in row 4 of the feature column; mean is in row 3 of the feature column
        formula = f"={cl(fcol)}$4*({cl(fcol)}{r}-{cl(fcol)}$3)"
        ws.cell(r, ccol, formula).number_format = "0.000000"

    # Summary block
    contrib_first = cl(COL_CONTRIB_START); contrib_last = cl(COL_CONTRIB_START + n_feat - 1)
    SUM_LOGIT_COL = COL_SUMMARY_START
    LOGIT_Z_COL = COL_SUMMARY_START + 1
    SMM_COL = COL_SUMMARY_START + 2
    CPR_COL = COL_SUMMARY_START + 3
    ws.cell(r, SUM_LOGIT_COL, f"=SUM({contrib_first}{r}:{contrib_last}{r})").number_format = "0.0000"
    ws.cell(r, LOGIT_Z_COL, f"=INTERCEPT_SCALED+{cl(SUM_LOGIT_COL)}{r}").number_format = "0.0000"
    ws.cell(r, SMM_COL, f"=1/(1+EXP(-{cl(LOGIT_Z_COL)}{r}))").number_format = "0.00000000"
    ws.cell(r, CPR_COL, f"=1-(1-{cl(SMM_COL)}{r})^12").number_format = "0.0000%"

    # Attribution columns: (contrib_j / total_logit_dev) * (pred_CPR - baseline_CPR)
    for k, f in enumerate(feature_list):
        ccol = CONTRIB_COL[f]
        acol = ATTR_COL[f]
        formula = (f"=IFERROR({cl(ccol)}{r}/{cl(SUM_LOGIT_COL)}{r}"
                   f"*({cl(CPR_COL)}{r}-BASELINE_CPR),0)")
        ws.cell(r, acol, formula).number_format = "0.0000%"

    # Sanity: baseline + sum(attr) - pred_cpr
    attr_first = cl(COL_ATTR_START); attr_last = cl(COL_ATTR_START + n_feat - 1)
    ws.cell(r, COL_SANITY, f"=BASELINE_CPR+SUM({attr_first}{r}:{attr_last}{r})-{cl(CPR_COL)}{r}").number_format = "0.000000%"

# Column widths (rough)
for j in range(1, COL_SANITY + 1):
    if j <= n_id:
        ws.column_dimensions[cl(j)].width = 22 if j == ID_LOAN else 16
    elif j < COL_FEAT_START:
        ws.column_dimensions[cl(j)].width = 12
    elif j < COL_CONTRIB_START:
        ws.column_dimensions[cl(j)].width = 14
    elif j < COL_SUMMARY_START:
        ws.column_dimensions[cl(j)].width = 14
    elif j < COL_ATTR_START:
        ws.column_dimensions[cl(j)].width = 12
    else:
        ws.column_dimensions[cl(j)].width = 14
ws.freeze_panes = ws.cell(DATA_START_ROW, COL_FEAT_START).coordinate

# ============================================================
# Sheet 4: Deal_Aggregator
# ============================================================
ws = wb.create_sheet("Deal_Aggregator")
ws.cell(1, 1, "Deal-level Rollup (UPB-weighted)").font = Font(bold=True, size=14)
ws.cell(2, 1, "Deal_ID is sourced from Loan_Snapshot column B. Edit there to regroup.").font = Font(italic=True, color="6B7280")

# Headers
agg_hdrs = ["Deal_ID", "n_loans", "total_UPB", "wavg_pred_CPR"]
agg_hdrs += [f"attr_{f}" for f in feature_list]
agg_hdrs += ["sanity_check"]
for j, h in enumerate(agg_hdrs, 1):
    c = ws.cell(4, j, h); c.fill = HEADER_FILL; c.font = WHITE; c.alignment = CENTER

# Row references in Loan_Snapshot
LS = "Loan_Snapshot"
LS_DEAL = f"{LS}!$B${DATA_START_ROW}:$B${DATA_START_ROW + N_LOANS - 1}"
LS_UPB = f"{LS}!${cl(IN_UPB)}${DATA_START_ROW}:${cl(IN_UPB)}${DATA_START_ROW + N_LOANS - 1}"
LS_CPR = f"{LS}!${cl(CPR_COL)}${DATA_START_ROW}:${cl(CPR_COL)}${DATA_START_ROW + N_LOANS - 1}"

# Default rollup: by deal_id (synthetic ISSUER_ERA grouping in starter data).
sample_with_deal = sample.copy()
sample_with_deal['_issuer_short'] = sample_with_deal['issuer_name'].fillna('').astype(str).str.split().str[0].str.replace(',', '').str.upper().str[:12]
sample_with_deal['_era'] = np.where(sample_with_deal['vintage_year']>=2021, 'POST21', 'PRE21')
sample_with_deal['_deal'] = sample_with_deal['_issuer_short'] + '_' + sample_with_deal['_era']
deal_counts = sample_with_deal['_deal'].value_counts()
deal_list = deal_counts.index.tolist()

agg_start_row = 5
for i, deal in enumerate(deal_list):
    r = agg_start_row + i
    ws.cell(r, 1, str(deal)).font = MONO
    # n_loans
    ws.cell(r, 2, f'=COUNTIF({LS_DEAL},$A{r})')
    # total_UPB
    ws.cell(r, 3, f'=SUMIF({LS_DEAL},$A{r},{LS_UPB})').number_format = "#,##0"
    # wavg_pred_CPR
    ws.cell(r, 4, f'=IFERROR(SUMPRODUCT(({LS_DEAL}=$A{r})*{LS_UPB}*{LS_CPR})/$C{r},0)').number_format = "0.0000%"
    # Per-feature attribution columns
    for k, f in enumerate(feature_list):
        attr_col_letter = cl(ATTR_COL[f])
        LS_ATTR = f"{LS}!${attr_col_letter}${DATA_START_ROW}:${attr_col_letter}${DATA_START_ROW + N_LOANS - 1}"
        ws.cell(r, 5 + k, f'=IFERROR(SUMPRODUCT(({LS_DEAL}=$A{r})*{LS_UPB}*{LS_ATTR})/$C{r},0)').number_format = "0.0000%"
    # Sanity check: baseline + sum(attr) - wavg_pred
    attr_first_col = cl(5)
    attr_last_col = cl(5 + n_feat - 1)
    ws.cell(r, 5 + n_feat, f"=BASELINE_CPR+SUM({attr_first_col}{r}:{attr_last_col}{r})-$D{r}").number_format = "0.000000%"

# A "TOTAL" row that aggregates ALL loans (deal_id wildcard)
total_row = agg_start_row + len(deal_list) + 2
ws.cell(total_row, 1, "ALL LOANS (snapshot total)").font = BOLD
ws.cell(total_row, 2, f'=COUNTA({LS}!$A${DATA_START_ROW}:$A${DATA_START_ROW + N_LOANS - 1})')
ws.cell(total_row, 3, f'=SUM({LS_UPB})').number_format = "#,##0"
ws.cell(total_row, 4, f'=SUMPRODUCT({LS_UPB}*{LS_CPR})/$C{total_row}').number_format = "0.0000%"
for k, f in enumerate(feature_list):
    attr_col_letter = cl(ATTR_COL[f])
    LS_ATTR = f"{LS}!${attr_col_letter}${DATA_START_ROW}:${attr_col_letter}${DATA_START_ROW + N_LOANS - 1}"
    ws.cell(total_row, 5 + k, f'=SUMPRODUCT({LS_UPB}*{LS_ATTR})/$C{total_row}').number_format = "0.0000%"
ws.cell(total_row, 5 + n_feat, f"=BASELINE_CPR+SUM({cl(5)}{total_row}:{cl(5+n_feat-1)}{total_row})-$D{total_row}").number_format = "0.000000%"
for j in range(1, 5 + n_feat + 1):
    ws.cell(total_row, j).fill = BASELINE_FILL

# Column widths
ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 14
for k in range(n_feat + 1):
    ws.column_dimensions[cl(5 + k)].width = 14
ws.freeze_panes = "B5"

# Save
wb.save(OUT_XLSX)
print(f"Wrote {OUT_XLSX}")
out_path = OUT_XLSX
print(f"  Loans in Loan_Snapshot: {N_LOANS}")
print(f"  Deals in Deal_Aggregator: {len(deal_list)} (+1 ALL row)")
print(f"  Total columns in Loan_Snapshot: {COL_SANITY}")