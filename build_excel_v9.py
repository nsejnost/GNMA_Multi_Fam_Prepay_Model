"""
V9 Excel calculator builder — Richards 5PL + per-cohort bounds + age-ramp.

Five tabs:
  1. README                  — orientation
  2. Cohort_Sigmoids         — per-cohort {floor, asymp, mid, slope, asym_factor,
                                age_knot_y0..y3, median_age} (one row per cohort)
  3. Structural_Mults        — V9-refit M_maturity, M_lockout, M_burnout
  4. Loan_Snapshot           — per-loan input → cohort lookup → Richards eval →
                                age_ramp eval → structural mults → CPR
  5. Benchmarks              — SanCap per-cohort grid

Math (V9):
    cohort_id  = lookup(program × purpose × size × age_bucket)
    cohort_smm = floor + asymp / (1 + EXP(-(net_refi_bps - mid)/slope))^(1/asym_factor)
                 ↑ Richards 5PL; reduces to 4PL when asym_factor=1
    age_ramp   = piecewise linear at knots [0,24,96,240]; normalized so ramp(median_age)=1
    pred_SMM   = MIN(cohort_smm × age_ramp × M_maturity × M_lockout × M_burnout, smm_cap)
    pred_CPR   = 1 - (1 - pred_SMM)^12
"""
import json, os, sys
from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

import cohort_v9 as c9

HERE = Path(__file__).resolve().parent
MODEL_JSON = HERE / "model_data_v9.json"
SAMPLE_PARQUET = HERE / "sample_loans.parquet"
OUT_XLSX = HERE / "V9_Excel_Calculator.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
SUBHEADER_FILL = PatternFill("solid", fgColor="374151")
PARAM_FILL = PatternFill("solid", fgColor="FEF3C7")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
MONO = Font(name="Consolas")
CENTER = Alignment(horizontal="center", vertical="center")
def cl(i): return get_column_letter(i)


def main():
    if not MODEL_JSON.exists():
        sys.exit(f"ERROR: {MODEL_JSON} not found - run train_v9.py first")
    md = json.load(open(MODEL_JSON))
    cohort_table = md['cohort_table']
    structural   = md['structural_params']

    panel_env = os.environ.get('GNMA_PANEL_PARQUET')
    if panel_env and Path(panel_env).exists():
        sample = regenerate_sample(panel_env)
    elif SAMPLE_PARQUET.exists():
        sample = pd.read_parquet(SAMPLE_PARQUET).reset_index(drop=True)
    else:
        sys.exit(f"ERROR: neither GNMA_PANEL_PARQUET nor {SAMPLE_PARQUET}")
    sample = sample[sample['loan_age_months'].notna() & sample['vintage_year'].notna()
                    ].reset_index(drop=True)

    print(f"Building {OUT_XLSX.name}  test_AUC={md['metadata']['test_auc']:.4f}  "
          f"max_pred_CPR={md['metadata']['max_cpr_full_panel']:.1f}%")
    print(f"  Loans in sample: {len(sample)}")
    print(f"  Cohort sigmoids in table: {len(cohort_table)}")

    wb = Workbook()
    wb.remove(wb.active)

    build_readme(wb, md)
    build_cohort_sigmoids(wb, cohort_table, md.get('cohort_definitions', []))
    build_structural_mults(wb, structural)
    build_loan_snapshot(wb, sample, cohort_table, structural, md['metadata']['smm_cap'])
    build_benchmarks_tab(wb, md)

    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}  ({OUT_XLSX.stat().st_size / 1024:.1f} KB)")


def regenerate_sample(panel_path):
    print(f"  Reading panel: {panel_path}")
    df = pd.read_parquet(panel_path)
    df = df.sort_values(['loan_id', 'period'])
    df['itm'] = (df['refi_incentive_bps'] > 0).astype(int)
    df['cum_itm'] = df.groupby('loan_id')['itm'].cumsum()
    snap_period = sorted(df['period'].astype(str).unique())[-2]
    snap = df[(df['period'].astype(str) == snap_period)
              & (df['prepay_eligible'] == 1) & (df['has_valid_data'] == 1)].copy()
    buckets = []
    for issuer_pat, n in [('LUMENT', 35), ('BERKADIA', 30), ('WALKER', 25),
                          ('GREYSTONE', 25), ('MERCHANTS', 20), ('PNC', 20),
                          ('WELLS FARGO', 20), ('CAPITAL FUNDING', 20),
                          ('DWIGHT', 15), ('KEYBANK', 15), ('JLL', 10),
                          ('BELLWETHER', 8), ('PGIM', 5), ('NORTHMARQ', 5)]:
        m = snap['issuer_name'].fillna('').str.upper().str.contains(issuer_pat, na=False, regex=False)
        sub = snap[m]
        if len(sub) > 0:
            buckets.append(sub.sample(min(n, len(sub)), random_state=42))
    out = pd.concat(buckets).drop_duplicates('loan_id').reset_index(drop=True)
    out.to_parquet(SAMPLE_PARQUET)
    return out


def build_readme(wb, md):
    ws = wb.create_sheet("README")
    rows = [
        ("V9 Voluntary Prepayment Calculator (Richards 5PL + per-cohort bounds + native aging)", BOLD),
        ("", None),
        (f"Test AUC: {md['metadata']['test_auc']:.4f}   Test log-loss: {md['metadata']['test_log_loss']:.5f}", None),
        (f"Cohorts fit: {md['metadata']['n_cohorts_fit']} primary + {md['metadata']['n_fallback_levels']} fallback", None),
        (f"Training period: {md['metadata']['period_range'][0]} → {md['metadata']['period_range'][1]}", None),
        (f"Max predicted CPR over full panel: {md['metadata']['max_cpr_full_panel']:.2f}%", None),
        ("", None),
        ("V9 enhancements (vs V8.1)", BOLD),
        ("  R1. Richards 5-parameter S-curve replaces 4PL logistic — adds asym_factor", MONO),
        ("       S(x) = floor + asymp / (1 + EXP(-(x-mid)/slope))^(1/asym_factor)", MONO),
        ("  R2. Per-cohort min/max bounds informed by program class:", MONO),
        ("       538 (rural)        : ceiling ~18 CPR (asymp ≤ 0.018)", MONO),
        ("       232 (HC)           : ceiling ~70 CPR (asymp ≤ 0.12)", MONO),
        ("       223a7 (streamlined): ceiling ~99 CPR (asymp ≤ 0.30)", MONO),
        ("       223f (acquisition) : ceiling ~92 CPR (asymp ≤ 0.22)", MONO),
        ("  R3. Widened bounds on M_lockout (4→8 amplitude) and M_maturity (8→16)", MONO),
        ("  R5. Age_ramp anchor at each cohort's median age (Case B with cohort-specific a₀)", MONO),
        ("", None),
        ("Tabs", BOLD),
        ("  Cohort_Sigmoids   - per-cohort Richards params (floor/asymp/mid/slope/asym_factor)", None),
        ("                       + age_ramp 4 knot heights + median_age", None),
        ("  Structural_Mults  - V9-refit M_maturity, M_lockout, M_burnout (R3)", None),
        ("  Loan_Snapshot     - cohort assignment, Richards eval, age_ramp interp, capped pred_CPR", None),
        ("  Benchmarks        - SanCap reference grid mapped to V9 cohorts", None),
    ]
    for i, (text, font) in enumerate(rows, 1):
        c = ws.cell(i, 1, text)
        if font is not None: c.font = font
    ws.column_dimensions['A'].width = 110


def build_cohort_sigmoids(wb, cohort_table, definitions):
    ws = wb.create_sheet("Cohort_Sigmoids")
    ws.cell(1, 1, "V9 Per-Cohort Richards-Sigmoid + Age-Ramp Lookup Table").font = Font(bold=True, size=14)
    ws.cell(2, 1, "Cohort_id format: 'program|purpose|size|age_bucket'. Fallback ids use '|ALL'; 'GLOBAL' is universal.").font = Font(italic=True, color="6B7280")
    ws.cell(3, 1, "Richards: S(x) = floor + asymp / (1 + EXP(-(x-mid)/slope))^(1/asym_factor). asym_factor=1 → symmetric 4PL.").font = Font(italic=True, color="6B7280")
    ws.cell(4, 1, "Age_ramp normalized so ramp(median_age)=1.0; the cohort sigmoid params describe rate response at the cohort's typical age (Case B).").font = Font(italic=True, color="6B7280")

    headers = ["cohort_id", "floor", "asymp", "mid", "slope", "asym_factor",
               "ramp_y0", "ramp_y1", "ramp_y2", "ramp_y3", "median_age",
               "n_loans", "n_events", "fit_quality"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(6, j, h); c.fill = HEADER_FILL; c.font = WHITE; c.alignment = CENTER

    diag_by_id = {d['cohort_id']: d for d in definitions}
    keys = sorted(cohort_table.keys(), key=lambda k: (k.count('ALL'), k == 'GLOBAL', k))
    for i, cid in enumerate(keys):
        r = 7 + i
        p = cohort_table[cid]
        ws.cell(r, 1, cid).font = MONO
        ws.cell(r, 2, p['floor']).number_format = "0.0000000"
        ws.cell(r, 3, p['asymp']).number_format = "0.0000000"
        ws.cell(r, 4, p['mid']).number_format = "0.00"
        ws.cell(r, 5, p['slope']).number_format = "0.00"
        ws.cell(r, 6, p.get('asym_factor', 1.0)).number_format = "0.0000"
        knots_y = p.get('age_knots_y', [1.0, 1.0, 1.0, 1.0])
        for k_i in range(4):
            ws.cell(r, 7 + k_i, knots_y[k_i]).number_format = "0.0000"
        ws.cell(r, 11, p.get('median_age', 96.0)).number_format = "0.0"
        d = diag_by_id.get(cid, {})
        ws.cell(r, 12, d.get('n', '')).number_format = "#,##0"
        ws.cell(r, 13, d.get('n_events', ''))
        ws.cell(r, 14, d.get('fit_quality', ''))

    end_row = 6 + len(keys)
    # Lookup range covers all 10 returnable columns (cohort_id + 5 sigmoid + 4 age_knots)
    table_ref = f"Cohort_Sigmoids!$A$7:$J${end_row}"
    wb.defined_names['V9_COHORT_TABLE'] = DefinedName('V9_COHORT_TABLE', attr_text=table_ref)

    for col, w in zip(range(1, 15), [34, 12, 12, 10, 10, 12, 11, 11, 11, 11, 12, 12, 10, 14]):
        ws.column_dimensions[cl(col)].width = w
    ws.freeze_panes = "A7"


def build_structural_mults(wb, structural):
    """V9-refit structural multipliers (R3 widened bounds: M_lockout up to 8,
    M_maturity up to 16). Re-fit on V9 cohort residuals."""
    ws = wb.create_sheet("Structural_Mults")
    ws.cell(1, 1, "V9 Structural Multipliers (R3: widened bounds, refit on V9 residuals)").font = Font(bold=True, size=14)

    row = 3
    def write(label, value, name, fmt="0.000000"):
        nonlocal row
        ws.cell(row, 1, label).font = MONO
        c = ws.cell(row, 2, value)
        c.font = MONO; c.fill = PARAM_FILL; c.number_format = fmt
        wb.defined_names[name] = DefinedName(name, attr_text=f"Structural_Mults!$B${row}")
        row += 1

    if structural.get('M_maturity'):
        ws.cell(row, 1, "M_maturity (1 + amp×max(0, (cutoff-mtm)/cutoff))   [R3: bound widened to amp ≤ 16]").font = BOLD
        row += 1
        write("  amplitude", structural['M_maturity']['amplitude'], 'V9_MTY_AMP', "0.0000")
        write("  cutoff",    structural['M_maturity']['cutoff'],    'V9_MTY_CUTOFF', "0.00")
        row += 1

    if structural.get('M_lockout'):
        ws.cell(row, 1, "M_lockout (1 + amp×exp(-msle/tau))   [R3: bound widened to amp ≤ 8]").font = BOLD
        row += 1
        write("  amplitude", structural['M_lockout']['amplitude'], 'V9_LOCKOUT_AMP', "0.0000")
        write("  tau",       structural['M_lockout']['tau'],       'V9_LOCKOUT_TAU', "0.0000")
        row += 1

    if structural.get('M_burnout'):
        ws.cell(row, 1, "M_burnout (max(floor, 1 - slope×burn_ratio); slope < 0 = running-ITM amplifier)").font = BOLD
        row += 1
        write("  floor", structural['M_burnout']['floor'], 'V9_BURN_FLOOR', "0.0000")
        write("  slope", structural['M_burnout']['slope'], 'V9_BURN_SLOPE', "0.0000")

    ws.column_dimensions['A'].width = 70
    ws.column_dimensions['B'].width = 14


def build_loan_snapshot(wb, sample, cohort_table, structural, smm_cap):
    ws = wb.create_sheet("Loan_Snapshot")

    headers = [
        ("ID", ["loan_id", "fha_category", "loan_purpose", "issuer_name",
                "vintage_year", "period"]),
        ("Inputs", ["upb", "loan_rate_pct", "plc_rate_bps", "prepay_penalty_points",
                    "loan_age_months", "loan_maturity_yyyymm", "lockout_end_yyyymm",
                    "cum_itm", "burn_ratio_in"]),
        ("Derived (cohort_id)", ["program_key", "purpose_key", "size_bucket",
                                  "age_bucket", "cohort_id",
                                  "gross_refi_bps", "net_refi_bps"]),
        ("Cohort sigmoid (Richards)", ["floor", "asymp", "mid", "slope", "asym_factor",
                                         "cohort_smm_anchor"]),
        ("Cohort age_ramp", ["ramp_y0", "ramp_y1", "ramp_y2", "ramp_y3", "age_ramp"]),
        ("Structural mults (V9 refit)", ["M_maturity", "M_lockout", "M_burnout"]),
        ("Output", ["pred_SMM", "pred_CPR"]),
    ]
    flat = [(grp, c) for grp, cols in headers for c in cols]
    n = len(flat)
    col = 1
    for grp, cols in headers:
        ws.merge_cells(start_row=1, end_row=1, start_column=col, end_column=col + len(cols) - 1)
        cell = ws.cell(1, col, grp); cell.fill = SUBHEADER_FILL; cell.font = WHITE
        cell.alignment = CENTER
        col += len(cols)
    for j, (_grp, c) in enumerate(flat, 1):
        cell = ws.cell(2, j, c); cell.fill = HEADER_FILL; cell.font = WHITE; cell.alignment = CENTER

    DATA_START = 3
    col_idx = {c: j + 1 for j, (_g, c) in enumerate(flat)}

    for i, s in sample.iterrows():
        r = DATA_START + i
        # ID
        ws.cell(r, col_idx['loan_id'], str(s['loan_id'])[-12:]).font = MONO
        ws.cell(r, col_idx['fha_category'], str(s.get('fha_category', '')))
        ws.cell(r, col_idx['loan_purpose'], str(s.get('loan_purpose', '')))
        ws.cell(r, col_idx['issuer_name'], str(s['issuer_name']))
        ws.cell(r, col_idx['vintage_year'], int(s['vintage_year']) if pd.notna(s['vintage_year']) else None)
        ws.cell(r, col_idx['period'], str(s['period']))

        # Inputs
        ws.cell(r, col_idx['upb'], float(s['upb'])).number_format = "#,##0"
        ws.cell(r, col_idx['loan_rate_pct'], float(s['loan_rate'])).number_format = "0.000"
        ws.cell(r, col_idx['plc_rate_bps'], float(s['plc_rate_bps'])).number_format = "0.0"
        ppp = float(s.get('prepay_penalty_points', 0)) if pd.notna(s.get('prepay_penalty_points', 0)) else 0
        ws.cell(r, col_idx['prepay_penalty_points'], ppp).number_format = "0.0"
        ws.cell(r, col_idx['loan_age_months'], float(s['loan_age_months']))
        mat_str = str(s.get('loan_maturity_date', '')).strip()
        ws.cell(r, col_idx['loan_maturity_yyyymm'],
                int(mat_str[:6]) if len(mat_str) >= 6 and mat_str[:6].isdigit() else '')
        lo_str = str(s.get('lockout_end_date', '')).strip()
        ws.cell(r, col_idx['lockout_end_yyyymm'],
                int(lo_str[:6]) if len(lo_str) >= 6 and lo_str[:6].isdigit() else '')
        ws.cell(r, col_idx['cum_itm'], int(s.get('cum_itm', 0)))
        age = max(1, float(s['loan_age_months']))
        br = min(1.0, max(0.0, float(s.get('cum_itm', 0)) / age))
        ws.cell(r, col_idx['burn_ratio_in'], br).number_format = "0.0000"

        # Derived (cohort_id)
        fha_col = cl(col_idx['fha_category'])
        ws.cell(r, col_idx['program_key'],
                f'=IF(ISNUMBER(SEARCH("232",{fha_col}{r})),"232",'
                f'IF(ISNUMBER(SEARCH("538",{fha_col}{r})),"538",'
                f'IF(ISNUMBER(SEARCH("223a7",{fha_col}{r})),"223a7",'
                f'IF(ISNUMBER(SEARCH("223f",{fha_col}{r})),"223f","OTHER"))))')
        purp_col = cl(col_idx['loan_purpose'])
        ws.cell(r, col_idx['purpose_key'],
                f'=IF(UPPER({purp_col}{r})="NC","NC","RP")')
        upb_col = cl(col_idx['upb'])
        ws.cell(r, col_idx['size_bucket'],
                f'=IF({upb_col}{r}<5000000,"small",IF({upb_col}{r}<25000000,"medium","large"))')
        age_col = cl(col_idx['loan_age_months'])
        ws.cell(r, col_idx['age_bucket'],
                f'=IF({age_col}{r}<36,"young",IF({age_col}{r}<120,"seasoned","aged"))')
        prog_col = cl(col_idx['program_key'])
        pur_col  = cl(col_idx['purpose_key'])
        sz_col   = cl(col_idx['size_bucket'])
        ab_col   = cl(col_idx['age_bucket'])
        ws.cell(r, col_idx['cohort_id'],
                f'={prog_col}{r}&"|"&{pur_col}{r}&"|"&{sz_col}{r}&"|"&{ab_col}{r}')
        ws.cell(r, col_idx['gross_refi_bps'],
                f"={cl(col_idx['loan_rate_pct'])}{r}*100-{cl(col_idx['plc_rate_bps'])}{r}"
                ).number_format = "0.0"
        ws.cell(r, col_idx['net_refi_bps'],
                f"={cl(col_idx['gross_refi_bps'])}{r}-12.5*({cl(col_idx['prepay_penalty_points'])}{r}+1)"
                ).number_format = "0.0"

        # Cohort lookup with 5-level VLOOKUP cascade
        coh_col = cl(col_idx['cohort_id'])
        fb1 = f'{prog_col}{r}&"|"&{pur_col}{r}&"|ALL|"&{ab_col}{r}'
        fb2 = f'{prog_col}{r}&"|ALL|ALL|"&{ab_col}{r}'
        fb3 = f'{prog_col}{r}&"|ALL|ALL|ALL"'

        # Lookup 5 sigmoid params (cols 2-6) + 4 age_knots (cols 7-10)
        # Note table has 10 returnable cols (column 1 is the cohort_id key)
        for col_name, table_col in [
                ('floor', 2), ('asymp', 3), ('mid', 4), ('slope', 5), ('asym_factor', 6),
                ('ramp_y0', 7), ('ramp_y1', 8), ('ramp_y2', 9), ('ramp_y3', 10)]:
            f_lookup = (
                f'=IFERROR(VLOOKUP({coh_col}{r},V9_COHORT_TABLE,{table_col},FALSE),'
                f'IFERROR(VLOOKUP({fb1},V9_COHORT_TABLE,{table_col},FALSE),'
                f'IFERROR(VLOOKUP({fb2},V9_COHORT_TABLE,{table_col},FALSE),'
                f'IFERROR(VLOOKUP({fb3},V9_COHORT_TABLE,{table_col},FALSE),'
                f'VLOOKUP("GLOBAL",V9_COHORT_TABLE,{table_col},FALSE)))))'
            )
            ws.cell(r, col_idx[col_name], f_lookup).number_format = "0.000000"

        # Richards 5PL: cohort_smm = floor + asymp / (1 + EXP(-(net-mid)/slope))^(1/asym_factor)
        net_col = cl(col_idx['net_refi_bps'])
        floor_col = cl(col_idx['floor']);   asymp_col = cl(col_idx['asymp'])
        mid_col   = cl(col_idx['mid']);     slope_col = cl(col_idx['slope'])
        asym_col  = cl(col_idx['asym_factor'])
        ws.cell(r, col_idx['cohort_smm_anchor'],
                f"={floor_col}{r}+{asymp_col}{r}*"
                f"(1+EXP(-({net_col}{r}-{mid_col}{r})/{slope_col}{r}))^(-1/MAX({asym_col}{r},0.001))"
                ).number_format = "0.000000"

        # Age ramp interp at loan_age_months across knots [0, 24, 96, 240]
        ry0 = cl(col_idx['ramp_y0']); ry1 = cl(col_idx['ramp_y1'])
        ry2 = cl(col_idx['ramp_y2']); ry3 = cl(col_idx['ramp_y3'])
        f_ramp = (
            f"=IF({age_col}{r}<=0,{ry0}{r},"
            f"IF({age_col}{r}>=240,{ry3}{r},"
            f"IF({age_col}{r}<=24,"
                f"{ry0}{r}+({age_col}{r}-0)/24*({ry1}{r}-{ry0}{r}),"
            f"IF({age_col}{r}<=96,"
                f"{ry1}{r}+({age_col}{r}-24)/72*({ry2}{r}-{ry1}{r}),"
                f"{ry2}{r}+({age_col}{r}-96)/144*({ry3}{r}-{ry2}{r})"
            f"))))"
        )
        ws.cell(r, col_idx['age_ramp'], f_ramp).number_format = "0.0000"

        # Structural multipliers (V9 refit; R3 widened bounds applied at fit time)
        period_col = cl(col_idx['period'])
        mat_col = cl(col_idx['loan_maturity_yyyymm'])
        mtm_expr = (f'IF(OR({mat_col}{r}="",NOT(ISNUMBER({mat_col}{r}))),360,'
                     f'MAX(0,MIN(480,'
                     f'(INT({mat_col}{r}/100)-INT(VALUE({period_col}{r})/100))*12'
                     f'+MOD({mat_col}{r},100)-MOD(VALUE({period_col}{r}),100))))')
        ws.cell(r, col_idx['M_maturity'],
                f"=IF({mtm_expr}>0,1+V9_MTY_AMP*MAX(0,(V9_MTY_CUTOFF-({mtm_expr}))/V9_MTY_CUTOFF),1)"
                ).number_format = "0.0000"

        lo_col = cl(col_idx['lockout_end_yyyymm'])
        msle_expr = (f'IF(OR({lo_col}{r}="",NOT(ISNUMBER({lo_col}{r}))),0,'
                     f'MAX(0,MIN(24,'
                     f'(INT(VALUE({period_col}{r})/100)-INT({lo_col}{r}/100))*12'
                     f'+MOD(VALUE({period_col}{r}),100)-MOD({lo_col}{r},100))))')
        ws.cell(r, col_idx['M_lockout'],
                f"=IF(AND({msle_expr}>0,{msle_expr}<=12),"
                f"1+V9_LOCKOUT_AMP*EXP(-({msle_expr})/V9_LOCKOUT_TAU),1)"
                ).number_format = "0.0000"

        br_col = cl(col_idx['burn_ratio_in'])
        ws.cell(r, col_idx['M_burnout'],
                f"=MAX(V9_BURN_FLOOR,1-V9_BURN_SLOPE*MIN(1,MAX(0,{br_col}{r})))"
                ).number_format = "0.0000"

        ws.cell(r, col_idx['pred_SMM'],
                f"=MIN({cl(col_idx['cohort_smm_anchor'])}{r}"
                f"*{cl(col_idx['age_ramp'])}{r}"
                f"*{cl(col_idx['M_maturity'])}{r}"
                f"*{cl(col_idx['M_lockout'])}{r}"
                f"*{cl(col_idx['M_burnout'])}{r},{smm_cap})"
                ).number_format = "0.000000"
        ws.cell(r, col_idx['pred_CPR'],
                f"=1-(1-{cl(col_idx['pred_SMM'])}{r})^12").number_format = "0.00%"

    for j in range(1, n + 1):
        ws.column_dimensions[cl(j)].width = 12 if j > len(headers[0][1]) else 16
    ws.freeze_panes = ws.cell(DATA_START, col_idx['cohort_id']).coordinate


def build_benchmarks_tab(wb, md):
    ws = wb.create_sheet("Benchmarks")
    ws.cell(1, 1, "SanCap per-cohort benchmarks vs V9").font = Font(bold=True, size=14)
    ws.cell(2, 1, "Each scenario maps to a V9 cohort_id; V9 prediction uses the Richards sigmoid + age_ramp + R3-refit structural mults.").font = Font(italic=True, color="6B7280")
    ws.cell(3, 1, "Persistent gaps still reflect the 2018-2026 vs 2014-2018 era mismatch (panel slower than SanCap reference). Informational only.").font = Font(italic=True, color="6B7280")

    grid = md.get('sancap_benchmarks', [])
    hdrs = ["Scenario", "Cohort", "SanCap CPR", "V9 pred CPR", "Δ (pp)", "Flag"]
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(4, j, h); c.fill = HEADER_FILL; c.font = WHITE; c.alignment = CENTER
    for i, r in enumerate(grid):
        row = 5 + i
        ws.cell(row, 1, r['scenario'])
        ws.cell(row, 2, r['cohort']).font = MONO
        ws.cell(row, 3, r['sancap_cpr']).number_format = "0.0"
        ws.cell(row, 4, r['v9_pred_cpr']).number_format = "0.00"
        ws.cell(row, 5, r['divergence_pp']).number_format = "0.00"
        flag_cell = ws.cell(row, 6, r['flag'])
        if r['flag'] == 'REVIEW':
            flag_cell.fill = PatternFill("solid", fgColor="FECACA")
            flag_cell.font = Font(bold=True, color="991B1B")
        else:
            flag_cell.fill = PatternFill("solid", fgColor="D1FAE5")
    for col, w in zip([1, 2, 3, 4, 5, 6], [38, 28, 14, 14, 12, 12]):
        ws.column_dimensions[cl(col)].width = w


if __name__ == '__main__':
    main()
