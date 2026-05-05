"""
V7 validation — three layers:

  Layer 1: Excel-vs-Python parity on 30 sample loans (max abs diff < 1e-6).
  Layer 2: Tail-blowup detector — score the FULL panel; assert max(pred_CPR) < 0.90
           (V7.1 raised the SMM cap from 0.10 to 0.15 — CPR cap from 72% to 84%).
  Layer 3: SanCap benchmark grid — flag any scenario diverging > 10pp.

Layers 1 & 2 are HARD pass/fail. Layer 3 is informational (V7 may legitimately
disagree with SanCap's 2014-2018 era benchmarks).
"""
import json, math, os, sys
from pathlib import Path
import numpy as np
import pandas as pd
import formulas
from openpyxl.utils import get_column_letter

import v7_multipliers as v7m

HERE = Path(__file__).resolve().parent
XLSX = HERE / "V7_Excel_Calculator.xlsx"
MODEL_JSON = HERE / "model_data_v7.json"
SAMPLE_PARQUET = HERE / "sample_loans.parquet"
PANEL_PATH = Path(os.environ.get('GNMA_PANEL_PARQUET',
                                  HERE / 'working' / 'gnma_mf_panel.parquet'))

if not all(p.exists() for p in (XLSX, MODEL_JSON, SAMPLE_PARQUET)):
    sys.exit(f"ERROR: missing one of {XLSX.name}, {MODEL_JSON.name}, {SAMPLE_PARQUET.name}")

md = json.load(open(MODEL_JSON))
params = md['multipliers']


# ============================================================================
# Layer 1: Excel-vs-Python parity
# ============================================================================
def derive_features_for_sample(s):
    """Compute V7 feature dict from one panel-row Series."""
    age = max(0.0, s['loan_age_months'] if pd.notna(s['loan_age_months']) else 0.0)
    upb = max(0.0, s['upb'] if pd.notna(s['upb']) else 0.0)
    period = int(s['period'])
    py, pm = period // 100, period % 100
    mat = str(s.get('loan_maturity_date', '')).strip()
    if len(mat) >= 6 and mat[:4].isdigit() and mat[4:6].isdigit():
        mtm = max(0, min(480, (int(mat[:4]) - py) * 12 + (int(mat[4:6]) - pm)))
    else:
        mtm = 360
    lo = str(s.get('lockout_end_date', '')).strip()
    if len(lo) >= 6 and lo[:4].isdigit() and lo[4:6].isdigit():
        msle = max(0, min(24, (py - int(lo[:4])) * 12 + (pm - int(lo[4:6]))))
    else:
        msle = 0
    cum_itm = float(s.get('cum_itm', 0))
    burn_ratio = min(1.0, max(0.0, cum_itm / max(1.0, age)))
    grf = float(s['loan_rate']) * 100 - float(s['plc_rate_bps'])
    ppp = float(s.get('prepay_penalty_points', 0) or 0)
    net_refi = grf - 12.5 * (ppp + 1)
    return {
        'loan_age_months':          np.array([age]),
        'net_refi_incentive_bps':   np.array([net_refi]),
        'gross_refi_incentive_bps': np.array([grf]),
        'prepay_penalty_points':    np.array([ppp]),
        'log_upb':                  np.array([math.log1p(upb)]),
        'fha_category':             np.array([str(s.get('fha_category', ''))]),
        'loan_purpose':             np.array([str(s.get('loan_purpose', ''))]),
        'months_since_lockout_end': np.array([float(msle)]),
        'months_to_maturity':       np.array([float(mtm)]),
        'burn_ratio':               np.array([burn_ratio]),
    }


def layer1():
    print(f"Layer 1 — loading {XLSX.name} via `formulas` (~30s)...")
    xl = formulas.ExcelModel().loads(str(XLSX)).finish().calculate()

    sample = pd.read_parquet(SAMPLE_PARQUET).reset_index(drop=True)
    sample = sample[sample['loan_age_months'].notna() & sample['vintage_year'].notna()].reset_index(drop=True)

    # Loan_Snapshot column layout (must match build_excel_v7.py)
    # V7.2: 8 multipliers + M_issuer = 9 mult cols; 3 derived (added net_refi_bps);
    # 8 attribution cols (no M_penalty).
    n_id, n_input, n_derived, n_mult, n_attr = 8, 11, 3, 9, 8
    COL_SUMM_START = 1 + n_id + n_input + n_derived + n_mult
    PRED_CPR_COL = COL_SUMM_START + 2
    SANITY_COL = COL_SUMM_START + 3 + n_attr
    PRED_CPR_LETTER = get_column_letter(PRED_CPR_COL)
    SANITY_LETTER = get_column_letter(SANITY_COL)
    DATA_START_ROW = 3
    print(f"  pred_CPR col = {PRED_CPR_LETTER}, sanity col = {SANITY_LETTER}")

    def lookup(col_letter, row):
        key = f"'[V7_Excel_Calculator.xlsx]LOAN_SNAPSHOT'!{col_letter}{row}"
        sol = xl.get(key)
        if sol is None:
            return None
        v = sol.value
        if hasattr(v, '__len__') and not isinstance(v, str):
            return v[0][0] if len(v) and len(v[0]) else None
        return v

    cpr_diffs, sanity_vals, sample_dump = [], [], []
    for i in range(min(30, len(sample))):
        r = DATA_START_ROW + i
        feats = derive_features_for_sample(sample.iloc[i])
        py_smm = float(v7m.predict_smm(feats, params)[0])
        py_cpr = (1 - (1 - py_smm) ** 12) * 100  # percent
        excel_cpr_pct = lookup(PRED_CPR_LETTER, r)
        if excel_cpr_pct is None:
            print(f"  row {r}: Excel CPR is None - skip")
            continue
        excel_cpr_pct = float(excel_cpr_pct) * 100  # Excel cell stored as decimal
        sanity = lookup(SANITY_LETTER, r)
        if sanity is not None:
            sanity_vals.append(float(sanity))
        diff = excel_cpr_pct - py_cpr
        cpr_diffs.append(diff)
        sample_dump.append((sample.iloc[i]['loan_id'], py_cpr, excel_cpr_pct, diff))

    diffs = np.array(cpr_diffs)
    print(f"  n={len(diffs)}  max|Excel-Python| pred_CPR = {np.max(np.abs(diffs)):.2e}%")
    print(f"  max|sanity| = {max((abs(v) for v in sanity_vals), default=0):.2e}")
    if len(diffs) and np.max(np.abs(diffs)) < 1e-6:
        print("  PASS")
        return True
    print("  FAIL — first divergent rows:")
    for loan_id, py, ex, d in sample_dump[:5]:
        print(f"    {loan_id[:30]:>30}  py={py:7.4f}%  excel={ex:7.4f}%  diff={d:+.4e}")
    return False


# ============================================================================
# Layer 2: Tail-blowup detector
# ============================================================================
def layer2():
    print(f"\nLayer 2 — full-panel tail-blowup detector...")
    print(f"  Loading panel: {PANEL_PATH}")
    df = pd.read_parquet(PANEL_PATH)
    last = sorted(df['period'].astype(str).unique())[-1]
    df = df[(df['prepay_eligible'] == 1)
            & (df['is_mature_loan'] != 1)
            & (df['period'].astype(str) != last)].copy()

    # Derive V7 features (matching train_v7.py)
    period_dt = pd.to_datetime(df['period'].astype(str), format='%Y%m')
    mat_dt = pd.to_datetime(df['loan_maturity_date'].astype(str).str.strip(),
                            format='%Y%m%d', errors='coerce')
    df['months_to_maturity'] = (
        ((mat_dt.dt.year - period_dt.dt.year) * 12
         + (mat_dt.dt.month - period_dt.dt.month))
        .fillna(360).clip(0, 480).astype(float)
    )
    lo_dt = pd.to_datetime(df['lockout_end_date'].astype(str).str.strip(),
                           format='%Y%m%d', errors='coerce')
    msle = ((period_dt.dt.year - lo_dt.dt.year) * 12
            + (period_dt.dt.month - lo_dt.dt.month))
    df['months_since_lockout_end'] = msle.where(msle.notna(), 0).clip(0, 24).astype(float)
    df['log_upb'] = np.log1p(df['upb'].fillna(0).clip(lower=0)).astype(float)
    df = df.sort_values(['loan_id', 'period']).reset_index(drop=True)
    df['_itm'] = (df['refi_incentive_bps'].fillna(-9999) > 0).astype(int)
    df['cum_itm'] = df.groupby('loan_id')['_itm'].cumsum()
    df['burn_ratio'] = (df['cum_itm'] / np.maximum(df['loan_age_months'].fillna(1), 1)).clip(0, 1).astype(float)
    df['gross_refi_incentive_bps'] = df['gross_refi_incentive_bps'].astype(float).fillna(0)
    df['prepay_penalty_points']    = df['prepay_penalty_points'].astype(float).fillna(0).clip(0, 10)
    df['loan_age_months']          = df['loan_age_months'].fillna(0).clip(lower=0).astype(float)
    df['fha_category']             = df['fha_category'].fillna('').astype(str)
    df['loan_purpose']             = df['loan_purpose'].fillna('').astype(str)
    if 'refi_incentive_bps' in df.columns:
        df['net_refi_incentive_bps'] = df['refi_incentive_bps'].astype(float).fillna(
            df['gross_refi_incentive_bps'] - 12.5 * (df['prepay_penalty_points'] + 1)
        )
    else:
        df['net_refi_incentive_bps'] = (df['gross_refi_incentive_bps']
                                         - 12.5 * (df['prepay_penalty_points'] + 1)).astype(float)

    feats = {
        'loan_age_months':          df['loan_age_months'].values,
        'net_refi_incentive_bps':   df['net_refi_incentive_bps'].values,
        'gross_refi_incentive_bps': df['gross_refi_incentive_bps'].values,
        'prepay_penalty_points':    df['prepay_penalty_points'].values,
        'log_upb':                  df['log_upb'].values,
        'fha_category':             df['fha_category'].values,
        'loan_purpose':             df['loan_purpose'].values,
        'months_since_lockout_end': df['months_since_lockout_end'].values,
        'months_to_maturity':       df['months_to_maturity'].values,
        'burn_ratio':               df['burn_ratio'].values,
    }
    smm = v7m.predict_smm(feats, params)
    cpr = (1 - (1 - smm) ** 12) * 100
    max_cpr = float(cpr.max())
    n_above_90 = int((cpr > 90).sum())
    n_above_75 = int((cpr > 75).sum())
    n_above_60 = int((cpr > 60).sum())
    print(f"  n loans scored: {len(cpr):,}")
    print(f"  max(pred_CPR) = {max_cpr:.2f}%")
    print(f"  count(pred_CPR > 90%) = {n_above_90:,}  > 75% = {n_above_75:,}  > 60% = {n_above_60:,}")
    # V7.1 threshold: 90% (cap is now CPR=84%, so anything > 90% is a real anomaly)
    if max_cpr < 90 and n_above_90 == 0:
        print("  PASS (no V6F-style blowups)")
        return True
    print("  FAIL — V7 predicts loan-months above the 90% CPR threshold")
    return False


# ============================================================================
# Layer 3: SanCap benchmark grid
# ============================================================================
def layer3():
    print(f"\nLayer 3 — SanCap benchmark grid (informational):")
    grid = md.get('sancap_benchmarks', [])
    n_review = sum(1 for r in grid if r['flag'] == 'REVIEW')
    print(f"  {n_review}/{len(grid)} scenarios diverge > 10pp")
    for r in grid:
        flag = '⚠' if r['flag'] == 'REVIEW' else ' '
        print(f"  {flag} {r['scenario']:<40}  SanCap={r['sancap_cpr']:>4} CPR  V7={r['v7_pred_cpr']:>5.1f} CPR  Δ={r['divergence_pp']:>5.1f}pp")
    return True   # informational only


if __name__ == '__main__':
    p1 = layer1()
    p2 = layer2()
    layer3()
    print()
    if p1 and p2:
        print("ALL HARD CHECKS PASS  (Layer 3 is informational)")
        sys.exit(0)
    print("FAILURES:", "Layer 1" if not p1 else "", "Layer 2" if not p2 else "")
    sys.exit(1)
