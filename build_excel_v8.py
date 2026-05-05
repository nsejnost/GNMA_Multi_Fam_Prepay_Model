"""
V8 Excel calculator builder — cohort-based S-curve architecture.

Six tabs:
  1. README              — orientation
  2. Cohort_Sigmoids     — per-cohort sigmoid parameters (named cells via INDEX/MATCH lookup)
  3. Structural_Mults    — V7's M_age/M_maturity/M_lockout/M_burnout (V8b reuses these)
  4. Loan_Snapshot       — per-loan inputs → cohort lookup → sigmoid → multipliers → CPR
  5. Deal_Aggregator     — UPB-weighted rollup
  6. Benchmarks          — SanCap per-cohort grid

Math form:
    cohort_id     = lookup(program × purpose × size_bucket)
    cohort_smm    = floor + asymp / (1 + EXP(-(net_refi_bps - mid)/slope))
    pred_SMM      = MIN(cohort_smm × M_age × M_maturity × M_lockout × M_burnout, smm_cap)
    pred_CPR      = 1 - (1 - pred_SMM)^12

Closed-form formulas throughout. INDEX/MATCH for the cohort lookup; named cells
for the structural multipliers reused from V7.
"""
import json, math, os, sys
from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

import cohort_v8 as c8
import v7_multipliers as v7m

HERE = Path(__file__).resolve().parent
MODEL_JSON = HERE / "model_data_v8.json"
SAMPLE_PARQUET = HERE / "sample_loans.parquet"
OUT_XLSX = HERE / "V8_Excel_Calculator.xlsx"


# ---------- Style helpers (match V6F/V7 conventions) ----------
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
SUBHEADER_FILL = PatternFill("solid", fgColor="374151")
PARAM_FILL = PatternFill("solid", fgColor="FEF3C7")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
MONO = Font(name="Consolas")
CENTER = Alignment(horizontal="center", vertical="center")
def cl(i): return get_column_letter(i)


def main():
    if not MODEL_JSON.exists():
        sys.exit(f"ERROR: {MODEL_JSON} not found - run train_v8.py first")
    md = json.load(open(MODEL_JSON))
    cohort_table = md['cohort_table']
    v7_struct    = md.get('v7_structural_params', {})

    panel_env = os.environ.get('GNMA_PANEL_PARQUET')
    if panel_env and Path(panel_env).exists():
        sample = regenerate_sample(panel_env)
    elif SAMPLE_PARQUET.exists():
        sample = pd.read_parquet(SAMPLE_PARQUET).reset_index(drop=True)
    else:
        sys.exit(f"ERROR: neither GNMA_PANEL_PARQUET nor {SAMPLE_PARQUET}")
    sample = sample[sample['loan_age_months'].notna() & sample['vintage_year'].notna()
                    ].reset_index(drop=True)

    print(f"Building {OUT_XLSX.name}  test_AUC={md['metadata']['test_auc']:.4f}  "
          f"max_pred_CPR={md['metadata']['max_cpr_full_panel']:.1f}%")
    print(f"  Loans in sample: {len(sample)}")
    print(f"  Cohort sigmoids in table: {len(cohort_table)}")

    wb = Workbook()
    wb.remove(wb.active)

    build_readme(wb, md)
    build_cohort_sigmoids(wb, cohort_table, md.get('cohort_definitions', []))
    build_structural_mults(wb, v7_struct)
    build_loan_snapshot(wb, sample, cohort_table, v7_struct, md['metadata']['smm_cap'])
    build_benchmarks_tab(wb, md)

    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX}  ({OUT_XLSX.stat().st_size / 1024:.1f} KB)")


def regenerate_sample(panel_path):
    """Same stratified-sample logic as build_excel_v7.py for parity."""
    print(f"  Reading panel: {panel_path}")
    df = pd.read_parquet(panel_path)
    df = df.sort_values(['loan_id', 'period'])
    df['itm'] = (df['refi_incentive_bps'] > 0).astype(int)
    df['cum_itm'] = df.groupby('loan_id')['itm'].cumsum()
    snap_period = sorted(df['period'].astype(str).unique())[-2]
    snap = df[(df['period'].astype(str) == snap_period)
              & (df['prepay_eligible'] == 1) & (df['has_valid_data'] == 1)].copy()
    buckets = []
    for issuer_pat, n in [
        ('LUMENT', 35), ('BERKADIA', 30), ('WALKER', 25), ('GREYSTONE', 25),
        ('MERCHANTS', 20), ('PNC', 20), ('WELLS FARGO', 20), ('CAPITAL FUNDING', 20),
        ('DWIGHT', 15), ('KEYBANK', 15), ('JLL', 10), ('BELLWETHER', 8),
        ('PGIM', 5), ('NORTHMARQ', 5),
    ]:
        m = snap['issuer_name'].fillna('').str.upper().str.contains(issuer_pat, na=False, regex=False)
        sub = snap[m]
        if len(sub) > 0:
            buckets.append(sub.sample(min(n, len(sub)), random_state=42))
    out = pd.concat(buckets).drop_duplicates('loan_id').reset_index(drop=True)
    out.to_parquet(SAMPLE_PARQUET)
    return out


# ---------------------------------------------------------------------------
# Sheet 1: README
# ---------------------------------------------------------------------------
def build_readme(wb, md):
    ws = wb.create_sheet("README")
    rows = [
        ("V8 Voluntary Prepayment Calculator (cohort-based S-curve)", BOLD),
        ("", None),
        (f"Test AUC: {md['metadata']['test_auc']:.4f}   "
         f"Test log-loss: {md['metadata']['test_log_loss']:.5f}   "
         f"Cohorts fit: {md['metadata']['n_cohorts_fit']}", None),
        (f"Training period: {md['metadata']['period_range'][0]} to {md['metadata']['period_range'][1]}", None),
        (f"Max predicted CPR over full panel: {md['metadata']['max_cpr_full_panel']:.2f}%   "
         f"(SMM cap: {md['metadata']['smm_cap']:.3f})", None),
        ("", None),
        ("Architecture", BOLD),
        ("  cohort_id   = program × purpose × size_bucket  (with hierarchical fallback)", MONO),
        ("  cohort_smm  = floor + asymp / (1 + EXP(-(net_refi_bps - mid)/slope))", MONO),
        ("  pred_SMM    = MIN(cohort_smm × M_age × M_maturity × M_lockout × M_burnout, smm_cap)", MONO),
        ("  pred_CPR    = 1 - (1 - pred_SMM)^12", MONO),
        ("", None),
        ("Tabs", BOLD),
        ("  Cohort_Sigmoids   - per-cohort fitted sigmoid params (cohort lookup table)", None),
        ("  Structural_Mults  - V7's age/maturity/lockout/burnout multipliers (reused as-is)", None),
        ("  Loan_Snapshot     - per-loan inputs, cohort lookup, sigmoid eval, capped CPR", None),
        ("  Benchmarks        - SanCap reference grid mapped to V8 cohorts", None),
        ("", None),
        ("Cohort definition (axes)", BOLD),
        (f"  program: {' / '.join(c8.PROGRAM_KEYS)}", MONO),
        (f"  purpose: {' / '.join(c8.PURPOSE_KEYS)}", MONO),
        (f"  size:    {' / '.join(c8.SIZE_BUCKET_KEYS)}  (breaks at $5M and $25M)", MONO),
    ]
    for i, (text, font) in enumerate(rows, 1):
        c = ws.cell(i, 1, text)
        if font is not None: c.font = font
    ws.column_dimensions['A'].width = 110


# ---------------------------------------------------------------------------
# Sheet 2: Cohort_Sigmoids — every cohort_id with its sigmoid params
# ---------------------------------------------------------------------------
def build_cohort_sigmoids(wb, cohort_table, definitions):
    """Cohort lookup table (one row per cohort, sigmoid + age_ramp params).

    V8.1: cohort_id is 4-part 'program|purpose|size|age_bucket'. Fallback ids
    use '|ALL' for dropped axes. Each row also stores the 4 age_ramp knot
    heights (knot_x positions are global constants in cohort_v8.py).
    """
    ws = wb.create_sheet("Cohort_Sigmoids")
    ws.cell(1, 1, "V8.1 Per-Cohort Sigmoid + Age-Ramp Lookup Table").font = Font(bold=True, size=14)
    ws.cell(2, 1, "Cohort_id format: 'program|purpose|size|age_bucket'. Fallback ids use '|ALL' for dropped axes; 'GLOBAL' is the universal fallback.").font = Font(italic=True, color="6B7280")
    ws.cell(3, 1, "Sigmoid evaluates SMM at age=96m (anchor); age_ramp scales it for other ages. Knot positions: 0/24/96/240 months.").font = Font(italic=True, color="6B7280")

    headers = ["cohort_id", "floor", "asymp", "mid", "slope",
               "ramp_y0", "ramp_y1", "ramp_y2_anchor", "ramp_y3",
               "n_loans", "n_events", "fit_quality"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(5, j, h); c.fill = HEADER_FILL; c.font = WHITE; c.alignment = CENTER

    diag_by_id = {d['cohort_id']: d for d in definitions}
    keys = sorted(cohort_table.keys(), key=lambda k: (k.count('ALL'), k == 'GLOBAL', k))
    for i, cid in enumerate(keys):
        r = 6 + i
        p = cohort_table[cid]
        ws.cell(r, 1, cid).font = MONO
        ws.cell(r, 2, p['floor']).number_format = "0.0000000"
        ws.cell(r, 3, p['asymp']).number_format = "0.0000000"
        ws.cell(r, 4, p['mid']).number_format = "0.00"
        ws.cell(r, 5, p['slope']).number_format = "0.00"
        knots_y = p.get('age_knots_y', [1.0, 1.0, 1.0, 1.0])
        for k_i in range(4):
            ws.cell(r, 6 + k_i, knots_y[k_i]).number_format = "0.0000"
        d = diag_by_id.get(cid, {})
        ws.cell(r, 10, d.get('n', '')).number_format = "#,##0"
        ws.cell(r, 11, d.get('n_events', ''))
        ws.cell(r, 12, d.get('fit_quality', ''))

    end_row = 5 + len(keys)
    # Range covers all 9 lookup columns (cohort_id + 4 sigmoid + 4 age_knots)
    table_ref = f"Cohort_Sigmoids!$A$6:$I${end_row}"
    wb.defined_names['V8_COHORT_TABLE'] = DefinedName('V8_COHORT_TABLE', attr_text=table_ref)

    for col, w in zip(range(1, 13), [34, 12, 12, 10, 10, 11, 11, 14, 11, 12, 10, 14]):
        ws.column_dimensions[cl(col)].width = w
    ws.freeze_panes = "A6"


# ---------------------------------------------------------------------------
# Sheet 3: Structural_Mults (V7's M_age, M_maturity, M_lockout, M_burnout)
# ---------------------------------------------------------------------------
def build_structural_mults(wb, v7_struct):
    """V8b reuses V7's age/maturity/lockout/burnout multipliers — they're cohort-
    invariant in both models and capture orthogonal signals.
    """
    ws = wb.create_sheet("Structural_Mults")
    ws.cell(1, 1, "V8b Structural Multipliers (reused from V7)").font = Font(bold=True, size=14)

    row = 3

    def write(label, value, name, fmt="0.000000"):
        nonlocal row
        ws.cell(row, 1, label).font = MONO
        c = ws.cell(row, 2, value)
        c.font = MONO; c.fill = PARAM_FILL; c.number_format = fmt
        wb.defined_names[name] = DefinedName(name, attr_text=f"Structural_Mults!$B${row}")
        row += 1

    if v7_struct.get('M_age'):
        ws.cell(row, 1, "M_age (piecewise linear, 6 knots)").font = BOLD; row += 1
        for i, (x, y) in enumerate(zip(v7_struct['M_age']['knots_x'],
                                        v7_struct['M_age']['knots_y'])):
            write(f"  knot {i}: x = {x:.0f}m → y", y, f'V8_AGE_KNOT_Y_{i}', "0.0000")
        ws.cell(row, 1, "  (knot_x values, frozen)").font = MONO; row += 1
        for i, x in enumerate(v7_struct['M_age']['knots_x']):
            ws.cell(row, 1, f"    knot_x_{i}").font = MONO
            c = ws.cell(row, 2, x); c.font = MONO
            wb.defined_names[f'V8_AGE_KNOT_X_{i}'] = DefinedName(
                f'V8_AGE_KNOT_X_{i}', attr_text=f"Structural_Mults!$B${row}")
            row += 1
        row += 1

    if v7_struct.get('M_maturity'):
        ws.cell(row, 1, "M_maturity (1 + amp×max(0, (cutoff-mtm)/cutoff))").font = BOLD; row += 1
        write("  amplitude", v7_struct['M_maturity']['amplitude'], 'V8_MTY_AMP', "0.0000")
        write("  cutoff",    v7_struct['M_maturity']['cutoff'],    'V8_MTY_CUTOFF', "0.00")
        row += 1

    if v7_struct.get('M_lockout'):
        ws.cell(row, 1, "M_lockout (1 + amp×exp(-msle/tau) for msle ∈ (0,12])").font = BOLD; row += 1
        write("  amplitude", v7_struct['M_lockout']['amplitude'], 'V8_LOCKOUT_AMP', "0.0000")
        write("  tau",       v7_struct['M_lockout']['tau'],       'V8_LOCKOUT_TAU', "0.0000")
        row += 1

    if v7_struct.get('M_burnout'):
        ws.cell(row, 1, "M_burnout (max(floor, 1 - slope×burn_ratio))").font = BOLD; row += 1
        write("  floor", v7_struct['M_burnout']['floor'], 'V8_BURN_FLOOR', "0.0000")
        write("  slope", v7_struct['M_burnout']['slope'], 'V8_BURN_SLOPE', "0.0000")

    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 14


# ---------------------------------------------------------------------------
# Sheet 4: Loan_Snapshot
# ---------------------------------------------------------------------------
def build_loan_snapshot(wb, sample, cohort_table, v7_struct, smm_cap):
    ws = wb.create_sheet("Loan_Snapshot")

    headers = [
        ("ID", ["loan_id", "fha_category", "loan_purpose", "issuer_name", "vintage_year", "period"]),
        ("Inputs", ["upb", "loan_rate_pct", "plc_rate_bps", "prepay_penalty_points",
                    "loan_age_months", "loan_maturity_yyyymm", "lockout_end_yyyymm",
                    "cum_itm", "burn_ratio_in"]),
        ("Derived (cohort id)", ["program_key", "purpose_key", "size_bucket",
                                  "age_bucket", "cohort_id",
                                  "gross_refi_bps", "net_refi_bps"]),
        ("Cohort sigmoid", ["floor", "asymp", "mid", "slope", "cohort_smm_anchor"]),
        ("Cohort age_ramp", ["ramp_y0", "ramp_y1", "ramp_y2", "ramp_y3", "age_ramp"]),
        ("Structural mults", ["M_maturity", "M_lockout", "M_burnout"]),
        ("Output", ["pred_SMM", "pred_CPR"]),
    ]
    flat = [(grp, c) for grp, cols in headers for c in cols]
    n = len(flat)
    # Banner row
    col = 1
    for grp, cols in headers:
        ws.merge_cells(start_row=1, end_row=1, start_column=col, end_column=col + len(cols) - 1)
        cell = ws.cell(1, col, grp); cell.fill = SUBHEADER_FILL; cell.font = WHITE
        cell.alignment = CENTER
        col += len(cols)
    # Header row
    for j, (_grp, c) in enumerate(flat, 1):
        cell = ws.cell(2, j, c); cell.fill = HEADER_FILL; cell.font = WHITE; cell.alignment = CENTER

    DATA_START = 3
    col_idx = {c: j + 1 for j, (_g, c) in enumerate(flat)}
    program_codes = c8.PROGRAM_KEYS

    for i, s in sample.iterrows():
        r = DATA_START + i
        # ID block
        ws.cell(r, col_idx['loan_id'], str(s['loan_id'])[-12:]).font = MONO
        ws.cell(r, col_idx['fha_category'], str(s.get('fha_category', '')))
        ws.cell(r, col_idx['loan_purpose'], str(s.get('loan_purpose', '')))
        ws.cell(r, col_idx['issuer_name'], str(s['issuer_name']))
        ws.cell(r, col_idx['vintage_year'],
                int(s['vintage_year']) if pd.notna(s['vintage_year']) else None)
        ws.cell(r, col_idx['period'], str(s['period']))

        # Inputs block
        ws.cell(r, col_idx['upb'], float(s['upb'])).number_format = "#,##0"
        ws.cell(r, col_idx['loan_rate_pct'], float(s['loan_rate'])).number_format = "0.000"
        ws.cell(r, col_idx['plc_rate_bps'], float(s['plc_rate_bps'])).number_format = "0.0"
        ppp = float(s.get('prepay_penalty_points', 0)) if pd.notna(s.get('prepay_penalty_points', 0)) else 0
        ws.cell(r, col_idx['prepay_penalty_points'], ppp).number_format = "0.0"
        ws.cell(r, col_idx['loan_age_months'], float(s['loan_age_months']))
        mat_str = str(s.get('loan_maturity_date', '')).strip()
        ws.cell(r, col_idx['loan_maturity_yyyymm'],
                int(mat_str[:6]) if len(mat_str) >= 6 and mat_str[:6].isdigit() else '')
        lo_str = str(s.get('lockout_end_date', '')).strip()
        ws.cell(r, col_idx['lockout_end_yyyymm'],
                int(lo_str[:6]) if len(lo_str) >= 6 and lo_str[:6].isdigit() else '')
        ws.cell(r, col_idx['cum_itm'], int(s.get('cum_itm', 0)))
        age = max(1, float(s['loan_age_months']))
        br = min(1.0, max(0.0, float(s.get('cum_itm', 0)) / age))
        ws.cell(r, col_idx['burn_ratio_in'], br).number_format = "0.0000"

        # Derived block — compute cohort_id via formulas
        # program_key: nested IF on fha_category substring matches
        fha_col = cl(col_idx['fha_category'])
        f_program = (
            f'=IF(ISNUMBER(SEARCH("232",{fha_col}{r})),"232",'
            f'IF(ISNUMBER(SEARCH("538",{fha_col}{r})),"538",'
            f'IF(ISNUMBER(SEARCH("223a7",{fha_col}{r})),"223a7",'
            f'IF(ISNUMBER(SEARCH("223f",{fha_col}{r})),"223f","OTHER"))))'
        )
        ws.cell(r, col_idx['program_key'], f_program)

        # purpose_key: NC vs RP (binary)
        purp_col = cl(col_idx['loan_purpose'])
        ws.cell(r, col_idx['purpose_key'],
                f'=IF(UPPER({purp_col}{r})="NC","NC","RP")')

        # size_bucket: small / medium / large from upb
        upb_col = cl(col_idx['upb'])
        ws.cell(r, col_idx['size_bucket'],
                f'=IF({upb_col}{r}<5000000,"small",IF({upb_col}{r}<25000000,"medium","large"))')

        # age_bucket: young / seasoned / aged from loan_age_months
        age_col = cl(col_idx['loan_age_months'])
        ws.cell(r, col_idx['age_bucket'],
                f'=IF({age_col}{r}<36,"young",IF({age_col}{r}<120,"seasoned","aged"))')

        # cohort_id = program | purpose | size | age_bucket
        prog_col = cl(col_idx['program_key'])
        pur_col  = cl(col_idx['purpose_key'])
        sz_col   = cl(col_idx['size_bucket'])
        ab_col   = cl(col_idx['age_bucket'])
        ws.cell(r, col_idx['cohort_id'],
                f'={prog_col}{r}&"|"&{pur_col}{r}&"|"&{sz_col}{r}&"|"&{ab_col}{r}')

        # gross / net refi bps
        ws.cell(r, col_idx['gross_refi_bps'],
                f"={cl(col_idx['loan_rate_pct'])}{r}*100-{cl(col_idx['plc_rate_bps'])}{r}"
                ).number_format = "0.0"
        ws.cell(r, col_idx['net_refi_bps'],
                f"={cl(col_idx['gross_refi_bps'])}{r}-12.5*({cl(col_idx['prepay_penalty_points'])}{r}+1)"
                ).number_format = "0.0"

        # Cohort lookup via VLOOKUP cascade — 5 fallback levels:
        # 0: program × purpose × size × age   (full cohort_id)
        # 1: drop size → program × purpose × ALL × age
        # 2: drop purpose → program × ALL × ALL × age
        # 3: drop age → program × ALL × ALL × ALL
        # 4: GLOBAL
        coh_col = cl(col_idx['cohort_id'])
        fb1 = f'{prog_col}{r}&"|"&{pur_col}{r}&"|ALL|"&{ab_col}{r}'
        fb2 = f'{prog_col}{r}&"|ALL|ALL|"&{ab_col}{r}'
        fb3 = f'{prog_col}{r}&"|ALL|ALL|ALL"'

        # Lookup the 4 sigmoid params (cols 2-5) and 4 age_ramp knots (cols 6-9).
        for col_name, table_col in [('floor', 2), ('asymp', 3), ('mid', 4), ('slope', 5),
                                      ('ramp_y0', 6), ('ramp_y1', 7),
                                      ('ramp_y2', 8), ('ramp_y3', 9)]:
            f_lookup = (
                f'=IFERROR(VLOOKUP({coh_col}{r},V8_COHORT_TABLE,{table_col},FALSE),'
                f'IFERROR(VLOOKUP({fb1},V8_COHORT_TABLE,{table_col},FALSE),'
                f'IFERROR(VLOOKUP({fb2},V8_COHORT_TABLE,{table_col},FALSE),'
                f'IFERROR(VLOOKUP({fb3},V8_COHORT_TABLE,{table_col},FALSE),'
                f'VLOOKUP("GLOBAL",V8_COHORT_TABLE,{table_col},FALSE)))))'
            )
            ws.cell(r, col_idx[col_name], f_lookup).number_format = "0.000000"

        # cohort_smm_anchor = sigmoid eval at anchor age (96m) — does NOT include ramp
        net_col = cl(col_idx['net_refi_bps'])
        floor_col = cl(col_idx['floor']);   asymp_col = cl(col_idx['asymp'])
        mid_col   = cl(col_idx['mid']);     slope_col = cl(col_idx['slope'])
        ws.cell(r, col_idx['cohort_smm_anchor'],
                f"={floor_col}{r}+{asymp_col}{r}/(1+EXP(-({net_col}{r}-{mid_col}{r})/{slope_col}{r}))"
                ).number_format = "0.000000"

        # age_ramp = piecewise linear interp of (knot_x=[0,24,96,240], knot_y=lookup)
        # at loan_age_months. 4 segments, with extrapolation flat outside [0, 240].
        ry0 = cl(col_idx['ramp_y0']); ry1 = cl(col_idx['ramp_y1'])
        ry2 = cl(col_idx['ramp_y2']); ry3 = cl(col_idx['ramp_y3'])
        f_ramp = (
            f"=IF({age_col}{r}<=0,{ry0}{r},"
            f"IF({age_col}{r}>=240,{ry3}{r},"
            f"IF({age_col}{r}<=24,"
                f"{ry0}{r}+({age_col}{r}-0)/24*({ry1}{r}-{ry0}{r}),"
            f"IF({age_col}{r}<=96,"
                f"{ry1}{r}+({age_col}{r}-24)/72*({ry2}{r}-{ry1}{r}),"
                f"{ry2}{r}+({age_col}{r}-96)/144*({ry3}{r}-{ry2}{r})"
            f"))))"
        )
        ws.cell(r, col_idx['age_ramp'], f_ramp).number_format = "0.0000"

        # Structural multipliers from V7 (M_maturity, M_lockout, M_burnout only;
        # M_age is REPLACED by the per-cohort age_ramp above)
        period_col = cl(col_idx['period'])
        mat_col = cl(col_idx['loan_maturity_yyyymm'])
        mtm_expr = (f'IF(OR({mat_col}{r}="",NOT(ISNUMBER({mat_col}{r}))),360,'
                     f'MAX(0,MIN(480,'
                     f'(INT({mat_col}{r}/100)-INT(VALUE({period_col}{r})/100))*12'
                     f'+MOD({mat_col}{r},100)-MOD(VALUE({period_col}{r}),100))))')
        ws.cell(r, col_idx['M_maturity'],
                f"=IF({mtm_expr}>0,1+V8_MTY_AMP*MAX(0,(V8_MTY_CUTOFF-({mtm_expr}))/V8_MTY_CUTOFF),1)"
                ).number_format = "0.0000"

        lo_col = cl(col_idx['lockout_end_yyyymm'])
        msle_expr = (f'IF(OR({lo_col}{r}="",NOT(ISNUMBER({lo_col}{r}))),0,'
                     f'MAX(0,MIN(24,'
                     f'(INT(VALUE({period_col}{r})/100)-INT({lo_col}{r}/100))*12'
                     f'+MOD(VALUE({period_col}{r}),100)-MOD({lo_col}{r},100))))')
        ws.cell(r, col_idx['M_lockout'],
                f"=IF(AND({msle_expr}>0,{msle_expr}<=12),"
                f"1+V8_LOCKOUT_AMP*EXP(-({msle_expr})/V8_LOCKOUT_TAU),1)"
                ).number_format = "0.0000"

        br_col = cl(col_idx['burn_ratio_in'])
        ws.cell(r, col_idx['M_burnout'],
                f"=MAX(V8_BURN_FLOOR,1-V8_BURN_SLOPE*MIN(1,MAX(0,{br_col}{r})))"
                ).number_format = "0.0000"

        # pred_SMM = MIN(cohort_sigmoid × age_ramp × structural mults, smm_cap)
        ws.cell(r, col_idx['pred_SMM'],
                f"=MIN({cl(col_idx['cohort_smm_anchor'])}{r}"
                f"*{cl(col_idx['age_ramp'])}{r}"
                f"*{cl(col_idx['M_maturity'])}{r}"
                f"*{cl(col_idx['M_lockout'])}{r}"
                f"*{cl(col_idx['M_burnout'])}{r},{smm_cap})"
                ).number_format = "0.000000"
        ws.cell(r, col_idx['pred_CPR'],
                f"=1-(1-{cl(col_idx['pred_SMM'])}{r})^12").number_format = "0.00%"

    # Column widths
    for j in range(1, n + 1):
        ws.column_dimensions[cl(j)].width = 12 if j > len(headers[0][1]) else 16
    ws.freeze_panes = ws.cell(DATA_START, col_idx['cohort_id']).coordinate


# ---------------------------------------------------------------------------
# Sheet 5: Benchmarks tab
# ---------------------------------------------------------------------------
def build_benchmarks_tab(wb, md):
    ws = wb.create_sheet("Benchmarks")
    ws.cell(1, 1, "SanCap per-cohort benchmarks vs V8").font = Font(bold=True, size=14)
    ws.cell(2, 1, "Each scenario maps to a V8 cohort_id; V8 prediction comes from that cohort's fitted sigmoid + V7 structural multipliers.").font = Font(italic=True, color="6B7280")
    ws.cell(3, 1, "Persistent gaps reflect the 2018-2026 vs 2014-2018 era mismatch (panel slower than SanCap reference). Informational, not pass/fail.").font = Font(italic=True, color="6B7280")

    grid = md.get('sancap_benchmarks', [])
    hdrs = ["Scenario", "Cohort", "SanCap CPR", "V8 pred CPR", "Δ (pp)", "Flag"]
    for j, h in enumerate(hdrs, 1):
        c = ws.cell(4, j, h); c.fill = HEADER_FILL; c.font = WHITE; c.alignment = CENTER
    for i, r in enumerate(grid):
        row = 5 + i
        ws.cell(row, 1, r['scenario'])
        ws.cell(row, 2, r['cohort']).font = MONO
        ws.cell(row, 3, r['sancap_cpr']).number_format = "0.0"
        ws.cell(row, 4, r['v8_pred_cpr']).number_format = "0.00"
        ws.cell(row, 5, r['divergence_pp']).number_format = "0.00"
        flag_cell = ws.cell(row, 6, r['flag'])
        if r['flag'] == 'REVIEW':
            flag_cell.fill = PatternFill("solid", fgColor="FECACA")
            flag_cell.font = Font(bold=True, color="991B1B")
        else:
            flag_cell.fill = PatternFill("solid", fgColor="D1FAE5")
    for col, w in zip([1, 2, 3, 4, 5, 6], [38, 26, 14, 14, 12, 12]):
        ws.column_dimensions[cl(col)].width = w


if __name__ == '__main__':
    main()
